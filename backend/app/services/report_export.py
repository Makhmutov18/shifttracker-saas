from __future__ import annotations

import calendar
import csv
import io
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models import Adjustment, AdjustmentType, PayrollRun, Shift, User, Venue
from app.utils import calculate_payout_total, safe_decimal


MONEY_FORMAT = '#,##0.00 [$₽-ru-RU]'
HOURS_FORMAT = '0.00'
PERCENT_FORMAT = '0.00%'
DATE_FORMAT = 'dd.mm.yyyy'
DATETIME_FORMAT = 'dd.mm.yyyy hh:mm'

MONTH_NAMES = (
    "январь", "февраль", "март", "апрель", "май", "июнь",
    "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь",
)
STATUS_LABELS = {
    "pending": "На подтверждении",
    "approved": "Утверждена",
    "rejected": "Отклонена",
}
PAY_MODEL_LABELS = {
    "hourly": "Почасовая",
    "fixed_shift": "Фикс за смену",
    "revenue": "Процент от выручки",
    "hybrid": "Ставка + процент",
}
RUN_STATUS_LABELS = {
    "draft": "Черновик",
    "finalized": "Зафиксирован",
    "paid": "Выплачен",
    "cancelled": "Отменён",
}
ADJUSTMENT_LABELS = {
    "bonus": "Бонус",
    "penalty": "Удержание",
}


def _enum_value(value: Any, fallback: str = "") -> str:
    raw = getattr(value, "value", value)
    return str(raw) if raw is not None else fallback


def safe_spreadsheet_text(value: Any, fallback: str = "") -> str:
    """Return user-controlled text without allowing spreadsheet formulas."""
    if value is None:
        return fallback
    text = str(value)
    if not text:
        return fallback
    if text[0] in ("=", "+", "-", "@"):
        return f"'{text}"
    return text


def _excel_datetime(value: Any) -> Any:
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _rate_text(user: User | None) -> str:
    if user is None:
        return "Не настроена"
    model = _enum_value(getattr(user, "pay_model", None), "hourly")
    rate = safe_decimal(getattr(user, "hourly_rate", None))
    percent = safe_decimal(getattr(user, "revenue_percentage", None))
    if model == "fixed_shift":
        return f"{rate:.2f} ₽/смена"
    if model == "revenue":
        return f"{percent:.2f}%"
    if model == "hybrid":
        return f"{rate:.2f} ₽/ч + {percent:.2f}%"
    return f"{rate:.2f} ₽/ч"


@dataclass
class ReportData:
    month: int
    year: int
    venue_name: str
    shifts: list[dict[str, Any]]
    employees: list[dict[str, Any]]
    adjustments: list[dict[str, Any]]
    payroll_runs: list[dict[str, Any]]
    overview: list[tuple[str, Any, str | None]]

    @property
    def period_label(self) -> str:
        return f"{MONTH_NAMES[self.month - 1].capitalize()} {self.year}"


def assemble_report_data(
    *,
    month: int,
    year: int,
    venue_name: str,
    shifts: Iterable[Shift],
    adjustments: Iterable[Adjustment],
    payroll_runs: Iterable[PayrollRun],
) -> ReportData:
    shift_rows: list[dict[str, Any]] = []
    adjustment_rows: list[dict[str, Any]] = []
    payroll_rows: list[dict[str, Any]] = []
    employee_totals: dict[uuid.UUID, dict[str, Any]] = {}

    status_counts = {"approved": 0, "pending": 0, "rejected": 0}
    approved_hours = Decimal("0")
    approved_amount = Decimal("0")
    approved_revenue = Decimal("0")
    cross_venue_count = 0

    for shift in shifts:
        employee = getattr(shift, "user", None)
        work_venue = getattr(shift, "venue", None)
        home_venue = getattr(employee, "venue", None) if employee is not None else None
        status = _enum_value(getattr(shift, "status", None), "pending")
        hours = safe_decimal(getattr(shift, "total_hours", None))
        salary = safe_decimal(getattr(shift, "salary_earned", None))
        revenue = safe_decimal(getattr(shift, "revenue", None))
        user_id = getattr(shift, "user_id", None)
        home_venue_id = getattr(employee, "venue_id", None) if employee is not None else None
        work_venue_id = getattr(shift, "venue_id", None)
        is_cross_venue = bool(home_venue_id and work_venue_id and home_venue_id != work_venue_id)

        if status in status_counts:
            status_counts[status] += 1
        if status == "approved":
            approved_hours += hours
            approved_amount += salary
            approved_revenue += revenue
            if is_cross_venue:
                cross_venue_count += 1

            if user_id is not None:
                totals = employee_totals.setdefault(
                    user_id,
                    {
                        "employee": safe_spreadsheet_text(getattr(employee, "name", None), "Сотрудник"),
                        "position": safe_spreadsheet_text(getattr(employee, "position", None), "Не указана"),
                        "home_venue": safe_spreadsheet_text(getattr(home_venue, "name", None), "Точка не указана"),
                        "approved_shifts": 0,
                        "hours": Decimal("0"),
                        "shift_amount": Decimal("0"),
                        "bonuses": Decimal("0"),
                        "deductions": Decimal("0"),
                        "cross_venue": 0,
                    },
                )
                totals["approved_shifts"] += 1
                totals["hours"] += hours
                totals["shift_amount"] += salary
                totals["cross_venue"] += int(is_cross_venue)

        shift_rows.append(
            {
                "date": getattr(shift, "date", None),
                "employee": safe_spreadsheet_text(getattr(employee, "name", None), "Сотрудник"),
                "position": safe_spreadsheet_text(getattr(employee, "position", None), "Не указана"),
                "home_venue": safe_spreadsheet_text(getattr(home_venue, "name", None), "Точка не указана"),
                "work_venue": safe_spreadsheet_text(getattr(work_venue, "name", None), "Точка не указана"),
                "start_time": getattr(shift, "start_time", None),
                "end_time": getattr(shift, "end_time", None),
                "hours": hours,
                "status": STATUS_LABELS.get(status, "Неизвестно"),
                "pay_model": PAY_MODEL_LABELS.get(
                    _enum_value(getattr(employee, "pay_model", None), "hourly"),
                    "Не настроена",
                ),
                "rate": _rate_text(employee),
                "revenue": revenue,
                # Historical source of truth. Never recalculate from current user rates.
                "salary": salary,
                "is_approved": status == "approved",
            }
        )

    bonus_total = Decimal("0")
    deduction_total = Decimal("0")
    for adjustment in adjustments:
        employee = getattr(adjustment, "user", None)
        venue = getattr(adjustment, "venue", None)
        creator = getattr(adjustment, "creator", None)
        adjustment_type = _enum_value(getattr(adjustment, "type", None), "penalty")
        amount = safe_decimal(getattr(adjustment, "amount", None))
        user_id = getattr(adjustment, "user_id", None)
        if adjustment_type == AdjustmentType.bonus.value:
            bonus_total += amount
        else:
            deduction_total += amount

        if user_id is not None:
            totals = employee_totals.setdefault(
                user_id,
                {
                    "employee": safe_spreadsheet_text(getattr(employee, "name", None), "Сотрудник"),
                    "position": safe_spreadsheet_text(getattr(employee, "position", None), "Не указана"),
                    "home_venue": safe_spreadsheet_text(
                        getattr(getattr(employee, "venue", None), "name", None), "Точка не указана"
                    ),
                    "approved_shifts": 0,
                    "hours": Decimal("0"),
                    "shift_amount": Decimal("0"),
                    "bonuses": Decimal("0"),
                    "deductions": Decimal("0"),
                    "cross_venue": 0,
                },
            )
            key = "bonuses" if adjustment_type == AdjustmentType.bonus.value else "deductions"
            totals[key] += amount

        adjustment_rows.append(
            {
                "date": getattr(adjustment, "created_at", None),
                "employee": safe_spreadsheet_text(getattr(employee, "name", None), "Сотрудник"),
                "venue": safe_spreadsheet_text(getattr(venue, "name", None), "Точка не указана"),
                "type": ADJUSTMENT_LABELS.get(adjustment_type, "Удержание"),
                "amount": amount,
                "reason": safe_spreadsheet_text(getattr(adjustment, "reason", None), "Не указана"),
                "creator": safe_spreadsheet_text(getattr(creator, "name", None), "Не указан"),
            }
        )

    employees = []
    for totals in sorted(employee_totals.values(), key=lambda item: item["employee"].lower()):
        totals["total"] = calculate_payout_total(
            totals["shift_amount"], totals["bonuses"], totals["deductions"]
        )
        employees.append(totals)

    fixed_total = Decimal("0")
    paid_total = Decimal("0")
    remaining_total = Decimal("0")
    snapshot_revenue = Decimal("0")
    for payroll_run in payroll_runs:
        status = _enum_value(getattr(payroll_run, "status", None), "draft")
        items = list(getattr(payroll_run, "items", None) or [])
        accrued = sum((safe_decimal(getattr(item, "final_amount", None)) for item in items), Decimal("0"))
        paid = sum((safe_decimal(getattr(item, "paid_amount", None)) for item in items), Decimal("0"))
        remaining = sum((safe_decimal(getattr(item, "remaining_amount", None)) for item in items), Decimal("0"))
        if not items:
            accrued = safe_decimal(getattr(payroll_run, "total_amount", None))
            paid = safe_decimal(getattr(payroll_run, "total_paid", None))
            remaining = max(Decimal("0"), accrued - paid)
        revenue = safe_decimal(getattr(payroll_run, "revenue_total", None))

        if status in ("finalized", "paid"):
            fixed_total += accrued
            paid_total += paid
            remaining_total += remaining
            snapshot_revenue += revenue

        payroll_rows.append(
            {
                "title": safe_spreadsheet_text(getattr(payroll_run, "title", None), "Расчёт выплаты"),
                "period_start": getattr(payroll_run, "period_start", None),
                "period_end": getattr(payroll_run, "period_end", None),
                "venue": safe_spreadsheet_text(
                    getattr(getattr(payroll_run, "venue", None), "name", None), "Все точки"
                ),
                "status": RUN_STATUS_LABELS.get(status, "Неизвестно"),
                "accrued": accrued,
                "paid": paid,
                "remaining": remaining,
                "revenue": revenue,
                "payroll_share": accrued / revenue if revenue > 0 else None,
                "completed_at": getattr(payroll_run, "paid_at", None)
                or getattr(payroll_run, "finalized_at", None),
            }
        )

    total_accrued = calculate_payout_total(approved_amount, bonus_total, deduction_total)
    revenue_for_share = snapshot_revenue if snapshot_revenue > 0 else approved_revenue
    share_amount = fixed_total if snapshot_revenue > 0 else total_accrued
    payroll_share = share_amount / revenue_for_share if revenue_for_share > 0 else None
    overview: list[tuple[str, Any, str | None]] = [
        ("Период", f"{MONTH_NAMES[month - 1].capitalize()} {year}", None),
        ("Точка", safe_spreadsheet_text(venue_name, "Все точки"), None),
        ("Утверждённые смены", status_counts["approved"], "integer"),
        ("На подтверждении", status_counts["pending"], "integer"),
        ("Отклонённые смены", status_counts["rejected"], "integer"),
        ("Утверждённые часы", approved_hours, "hours"),
        ("Начислено за смены", approved_amount, "money"),
        ("Бонусы", bonus_total, "money"),
        ("Удержания", deduction_total, "money"),
        ("Итого к выплате", total_accrued, "money"),
        ("Зафиксировано", fixed_total, "money"),
        ("Выплачено", paid_total, "money"),
        ("Осталось", remaining_total, "money"),
        ("Сотрудники", len(employee_totals), "integer"),
        ("Cross-venue смены", cross_venue_count, "integer"),
        ("Выручка", revenue_for_share, "money"),
        ("Доля ФОТ", payroll_share, "percent" if payroll_share is not None else None),
    ]

    return ReportData(
        month=month,
        year=year,
        venue_name=venue_name,
        shifts=sorted(shift_rows, key=lambda row: (row["date"] or date.min, row["employee"])),
        employees=employees,
        adjustments=sorted(adjustment_rows, key=lambda row: str(row["date"] or "")),
        payroll_runs=sorted(payroll_rows, key=lambda row: row["period_start"] or date.min),
        overview=overview,
    )


async def load_report_data(
    session: AsyncSession,
    *,
    month: int,
    year: int,
    venue_id: uuid.UUID | None,
) -> ReportData:
    period_start = date(year, month, 1)
    period_end = date(year, month, calendar.monthrange(year, month)[1])

    shifts_query = (
        select(Shift)
        .options(selectinload(Shift.user).selectinload(User.venue), selectinload(Shift.venue))
        .where(Shift.date >= period_start, Shift.date <= period_end)
        .order_by(Shift.date, Shift.start_time)
    )
    adjustments_query = (
        select(Adjustment)
        .options(
            selectinload(Adjustment.user).selectinload(User.venue),
            selectinload(Adjustment.venue),
            selectinload(Adjustment.creator),
        )
        .where(Adjustment.month == month, Adjustment.year == year)
        .order_by(Adjustment.created_at)
    )
    payroll_query = (
        select(PayrollRun)
        .options(selectinload(PayrollRun.venue), selectinload(PayrollRun.items))
        .where(PayrollRun.period_start <= period_end, PayrollRun.period_end >= period_start)
        .order_by(PayrollRun.period_start, PayrollRun.created_at)
    )
    if venue_id is not None:
        shifts_query = shifts_query.where(Shift.venue_id == venue_id)
        adjustments_query = adjustments_query.where(Adjustment.venue_id == venue_id)
        payroll_query = payroll_query.where(PayrollRun.venue_id == venue_id)

    shifts = list((await session.execute(shifts_query)).scalars().all())
    adjustments = list((await session.execute(adjustments_query)).scalars().all())
    payroll_runs = list((await session.execute(payroll_query)).scalars().all())
    venue = await session.get(Venue, venue_id) if venue_id is not None else None
    venue_name = safe_spreadsheet_text(getattr(venue, "name", None), "Все точки")

    return assemble_report_data(
        month=month,
        year=year,
        venue_name=venue_name,
        shifts=shifts,
        adjustments=adjustments,
        payroll_runs=payroll_runs,
    )


HEADER_FILL = PatternFill("solid", fgColor="2563EB")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="111827", bold=True, size=15)
SUBTITLE_FONT = Font(color="64748B", size=10)
TOTAL_FILL = PatternFill("solid", fgColor="E2E8F0")
THIN_BORDER = Border(bottom=Side(style="thin", color="CBD5E1"))


def _set_title(sheet, report: ReportData, columns: int) -> None:
    last_column = get_column_letter(columns)
    sheet.merge_cells(f"A1:{last_column}1")
    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A1"] = "Порядок.Смены — отчёт"
    sheet["A2"] = f"Период: {report.period_label} · Точка: {report.venue_name}"
    sheet["A1"].font = TITLE_FONT
    sheet["A2"].font = SUBTITLE_FONT
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet["A2"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 24


def _prepare_sheet(sheet, report: ReportData, headers: list[str]) -> None:
    _set_title(sheet, report, len(headers))
    for index, header in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.freeze_panes = "A5"
    sheet.sheet_view.showGridLines = False
    sheet.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.outlinePr.summaryBelow = True
    sheet.row_dimensions[4].height = 30


def _add_table(sheet, *, name: str, columns: int, last_row: int) -> None:
    if last_row < 5:
        return
    reference = f"A4:{get_column_letter(columns)}{last_row}"
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def _finish_sheet(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_empty(sheet, columns: int, message: str) -> None:
    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=columns)
    cell = sheet.cell(row=5, column=1, value=message)
    cell.font = Font(color="64748B", italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[5].height = 28


def _write_overview(workbook: Workbook, report: ReportData) -> None:
    sheet = workbook.active
    sheet.title = "Обзор"
    headers = ["Показатель", "Значение"]
    _prepare_sheet(sheet, report, headers)
    for row_number, (label, value, value_type) in enumerate(report.overview, 5):
        sheet.cell(row=row_number, column=1, value=label)
        cell = sheet.cell(row=row_number, column=2, value=value)
        if value_type == "money":
            cell.number_format = MONEY_FORMAT
        elif value_type == "hours":
            cell.number_format = HOURS_FORMAT
        elif value_type == "percent":
            cell.number_format = PERCENT_FORMAT
        elif value_type == "integer":
            cell.number_format = "0"
    _add_table(sheet, name="OverviewTable", columns=2, last_row=4 + len(report.overview))
    _finish_sheet(sheet, [28, 24])


def _write_shifts(workbook: Workbook, report: ReportData) -> None:
    sheet = workbook.create_sheet("Смены")
    headers = [
        "Дата", "Сотрудник", "Должность", "Основная точка", "Точка смены",
        "Начало", "Конец", "Часы", "Статус", "Модель оплаты", "Текущая ставка",
        "Выручка смены", "Начислено",
    ]
    _prepare_sheet(sheet, report, headers)
    if not report.shifts:
        _write_empty(sheet, len(headers), "За выбранный период смен нет")
    else:
        for row_number, item in enumerate(report.shifts, 5):
            values = [
                item["date"], item["employee"], item["position"], item["home_venue"],
                item["work_venue"], item["start_time"], item["end_time"], item["hours"],
                item["status"], item["pay_model"], item["rate"], item["revenue"], item["salary"],
            ]
            for column, value in enumerate(values, 1):
                sheet.cell(row=row_number, column=column, value=value)
            for column in (1,):
                sheet.cell(row=row_number, column=column).number_format = DATE_FORMAT
            for column in (6, 7):
                sheet.cell(row=row_number, column=column).number_format = "hh:mm"
            sheet.cell(row=row_number, column=8).number_format = HOURS_FORMAT
            for column in (12, 13):
                sheet.cell(row=row_number, column=column).number_format = MONEY_FORMAT
        _add_table(sheet, name="ShiftsTable", columns=len(headers), last_row=4 + len(report.shifts))
        total_row = 5 + len(report.shifts)
        approved = [row for row in report.shifts if row["is_approved"]]
        totals = [
            "Итого по утверждённым", "", "", "", "", "", "",
            sum((row["hours"] for row in approved), Decimal("0")), "", "", "",
            sum((row["revenue"] for row in approved), Decimal("0")),
            sum((row["salary"] for row in approved), Decimal("0")),
        ]
        for column, value in enumerate(totals, 1):
            cell = sheet.cell(row=total_row, column=column, value=value)
            cell.font = Font(bold=True)
            cell.fill = TOTAL_FILL
            cell.border = THIN_BORDER
        sheet.cell(total_row, 8).number_format = HOURS_FORMAT
        sheet.cell(total_row, 12).number_format = MONEY_FORMAT
        sheet.cell(total_row, 13).number_format = MONEY_FORMAT
    _finish_sheet(sheet, [13, 24, 20, 22, 22, 11, 11, 10, 18, 20, 20, 17, 17])


def _write_employees(workbook: Workbook, report: ReportData) -> None:
    sheet = workbook.create_sheet("Сотрудники")
    headers = [
        "Сотрудник", "Должность", "Основная точка", "Утверждённых смен", "Часы",
        "Начислено за смены", "Бонусы", "Удержания", "Итого", "Cross-venue смены",
    ]
    _prepare_sheet(sheet, report, headers)
    if not report.employees:
        _write_empty(sheet, len(headers), "За выбранный период начислений нет")
    else:
        for row_number, item in enumerate(report.employees, 5):
            values = [
                item["employee"], item["position"], item["home_venue"], item["approved_shifts"],
                item["hours"], item["shift_amount"], item["bonuses"], item["deductions"],
                item["total"], item["cross_venue"],
            ]
            for column, value in enumerate(values, 1):
                sheet.cell(row=row_number, column=column, value=value)
            sheet.cell(row_number, 5).number_format = HOURS_FORMAT
            for column in (6, 7, 8, 9):
                sheet.cell(row_number, column).number_format = MONEY_FORMAT
        _add_table(sheet, name="EmployeesTable", columns=len(headers), last_row=4 + len(report.employees))
        total_row = 5 + len(report.employees)
        sheet.cell(total_row, 1, "Итого")
        for column, key in ((4, "approved_shifts"), (5, "hours"), (6, "shift_amount"), (7, "bonuses"), (8, "deductions"), (9, "total"), (10, "cross_venue")):
            sheet.cell(total_row, column, sum((row[key] for row in report.employees), Decimal("0")))
        for cell in sheet[total_row]:
            cell.font = Font(bold=True)
            cell.fill = TOTAL_FILL
        sheet.cell(total_row, 5).number_format = HOURS_FORMAT
        for column in (6, 7, 8, 9):
            sheet.cell(total_row, column).number_format = MONEY_FORMAT
    _finish_sheet(sheet, [24, 20, 22, 18, 11, 20, 15, 15, 18, 18])


def _write_adjustments(workbook: Workbook, report: ReportData) -> None:
    sheet = workbook.create_sheet("Корректировки")
    headers = ["Дата", "Сотрудник", "Точка", "Тип", "Сумма", "Причина", "Кто добавил"]
    _prepare_sheet(sheet, report, headers)
    if not report.adjustments:
        _write_empty(sheet, len(headers), "За выбранный период корректировок нет")
    else:
        for row_number, item in enumerate(report.adjustments, 5):
            values = [_excel_datetime(item["date"]), item["employee"], item["venue"], item["type"], item["amount"], item["reason"], item["creator"]]
            for column, value in enumerate(values, 1):
                sheet.cell(row=row_number, column=column, value=value)
            sheet.cell(row_number, 1).number_format = DATETIME_FORMAT
            sheet.cell(row_number, 5).number_format = MONEY_FORMAT
        _add_table(sheet, name="AdjustmentsTable", columns=len(headers), last_row=4 + len(report.adjustments))
        total_row = 5 + len(report.adjustments)
        sheet.cell(total_row, 1, "Итого")
        sheet.cell(total_row, 5, sum((row["amount"] for row in report.adjustments), Decimal("0")))
        sheet.cell(total_row, 5).number_format = MONEY_FORMAT
        for cell in sheet[total_row]:
            cell.font = Font(bold=True)
            cell.fill = TOTAL_FILL
    _finish_sheet(sheet, [18, 24, 22, 15, 16, 42, 24])


def _write_payroll_runs(workbook: Workbook, report: ReportData) -> None:
    sheet = workbook.create_sheet("Расчёты и выплаты")
    headers = [
        "Название расчёта", "Начало периода", "Конец периода", "Точка", "Статус",
        "Начислено", "Выплачено", "Остаток", "Выручка", "Доля ФОТ", "Дата фиксации / выплаты",
    ]
    _prepare_sheet(sheet, report, headers)
    if not report.payroll_runs:
        _write_empty(sheet, len(headers), "За выбранный период расчётов нет")
    else:
        for row_number, item in enumerate(report.payroll_runs, 5):
            values = [
                item["title"], item["period_start"], item["period_end"], item["venue"], item["status"],
                item["accrued"], item["paid"], item["remaining"], item["revenue"],
                item["payroll_share"], _excel_datetime(item["completed_at"]),
            ]
            for column, value in enumerate(values, 1):
                sheet.cell(row=row_number, column=column, value=value)
            for column in (2, 3):
                sheet.cell(row_number, column).number_format = DATE_FORMAT
            for column in (6, 7, 8, 9):
                sheet.cell(row_number, column).number_format = MONEY_FORMAT
            sheet.cell(row_number, 10).number_format = PERCENT_FORMAT
            sheet.cell(row_number, 11).number_format = DATETIME_FORMAT
        _add_table(sheet, name="PayrollRunsTable", columns=len(headers), last_row=4 + len(report.payroll_runs))
        total_row = 5 + len(report.payroll_runs)
        sheet.cell(total_row, 1, "Итого")
        for column, key in ((6, "accrued"), (7, "paid"), (8, "remaining"), (9, "revenue")):
            sheet.cell(total_row, column, sum((row[key] for row in report.payroll_runs), Decimal("0")))
            sheet.cell(total_row, column).number_format = MONEY_FORMAT
        for cell in sheet[total_row]:
            cell.font = Font(bold=True)
            cell.fill = TOTAL_FILL
    _finish_sheet(sheet, [30, 16, 16, 22, 18, 17, 17, 17, 17, 13, 23])


def build_xlsx(report: ReportData) -> bytes:
    workbook = Workbook()
    _write_overview(workbook, report)
    _write_shifts(workbook, report)
    _write_employees(workbook, report)
    _write_adjustments(workbook, report)
    _write_payroll_runs(workbook, report)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_csv(report: ReportData) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output, delimiter=";", lineterminator="\r\n")
    writer.writerow([
        "Дата", "Сотрудник", "Должность", "Основная точка", "Точка смены",
        "Начало", "Конец", "Часы", "Статус", "Модель оплаты", "Ставка",
        "Выручка смены", "Начислено",
    ])
    for item in report.shifts:
        writer.writerow([
            item["date"].strftime("%d.%m.%Y") if item["date"] else "",
            item["employee"], item["position"], item["home_venue"], item["work_venue"],
            item["start_time"].strftime("%H:%M") if item["start_time"] else "",
            item["end_time"].strftime("%H:%M") if item["end_time"] else "",
            f'{item["hours"]:.2f}', item["status"], item["pay_model"], item["rate"],
            f'{item["revenue"]:.2f}', f'{item["salary"]:.2f}',
        ])
    return b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")
