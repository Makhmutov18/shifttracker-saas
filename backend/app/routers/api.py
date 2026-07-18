from fastapi import APIRouter, Depends, HTTPException, Header, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, and_, case
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import date, timedelta, datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
import io
import openpyxl
from urllib.parse import quote
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.database import get_session
from app.models import (
    User, Venue, Shift, Expense, UserRole, ShiftStatus, AuditLog, Adjustment, AdjustmentType,
    PayrollRun, PayrollRunItem, PayrollRunStatus, PayrollPayment,
    PayrollRunShiftSource, PayrollRunAdjustmentSource,
)
from app.permissions import has_permission
from app.schemas import (
    UserOut, VenueOut, ShiftCreate, ShiftOut, ShiftUpdate,
    ExpenseCreate, ExpenseOut, MonthlyStats,
    AuditLogOut, AdjustmentCreate, AdjustmentOut,
    PayrollSummaryOut, PayrollSummaryRow, PayrollPreviewOut, PayrollPreviewRow,
    PayrollRunCreate, PayrollRunRead, PayrollRunListItem, PayrollRunItemRead, PayrollPaymentRead,
    PayrollRunRevenueUpdate,
    PayrollPaymentCreate, PayrollPaymentResult, PersonalPayrollRunRead, PersonalPayrollPaymentRead,
    VenueStatsRow,
)
from app.auth import authenticate_request
from app.utils import (
    calculate_hours,
    calculate_salary,
    normalize_pay_model,
    safe_decimal,
    calculate_payout_total,
    calculate_payroll_share,
    safe_text,
    shift_status_label,
)
from app.notifications import notify_shift_approved, notify_shift_rejected, notify_bonus_added, notify_penalty_added

import uuid
import logging

router = APIRouter(prefix="/api", tags=["api"])
logger = logging.getLogger(__name__)


def _can_manage_all_venue_shifts(user: User) -> bool:
    return user.role in (UserRole.owner, UserRole.admin)


def _can_view_team_history_scope(user: User) -> bool:
    return (
        _can_manage_all_venue_shifts(user)
        or has_permission(user, "can_view_team_payroll")
        or has_permission(user, "can_approve_shifts")
        or has_permission(user, "can_edit_team_shifts")
    )


def _can_view_general_audit(user: User) -> bool:
    return user.role in (UserRole.owner, UserRole.admin) or has_permission(user, "can_manage_team")


def _can_view_payroll_runs(user: User) -> bool:
    return has_permission(user, "can_view_team_payroll")


def _payroll_run_is_visible(run: PayrollRun, user: User) -> bool:
    if _can_manage_all_venue_shifts(user):
        return True
    return run.venue_id is not None and run.venue_id == user.venue_id


def _can_send_shift_reminders(user: User) -> bool:
    return user.role in (UserRole.owner, UserRole.admin) or has_permission(user, "can_manage_team")


def _serialize_audit_logs(logs: list[AuditLog]) -> list[AuditLogOut]:
    return [
        AuditLogOut(
            id=log.id,
            user_id=log.user_id,
            target_user_id=log.target_user_id,
            action=log.action,
            entity_type=log.entity_type,
            entity_id=log.entity_id,
            old_value=log.old_value,
            new_value=log.new_value,
            created_at=log.created_at.isoformat() if log.created_at else "",
            user_name=log.user.name if log.user else None,
            target_user_name=log.target_user.name if log.target_user else None,
        )
        for log in logs
    ]


async def get_current_user(
    request: Request,
    init_data: str | None = Header(None, alias="X-Init-Data"),
    session: AsyncSession = Depends(get_session),
) -> User:
    """Dependency: validates Telegram initData or the secure web session."""
    return await authenticate_request(request, init_data, session)


async def _get_active_venue(session: AsyncSession, venue_id: uuid.UUID) -> Venue:
    result = await session.execute(
        select(Venue).where(Venue.id == venue_id, Venue.is_active == True)
    )
    venue = result.scalar_one_or_none()
    if venue is None:
        raise HTTPException(status_code=404, detail="Активная точка не найдена")
    return venue


async def _load_shift_response(session: AsyncSession, shift_id: uuid.UUID) -> Shift:
    result = await session.execute(
        select(Shift)
        .options(selectinload(Shift.user), selectinload(Shift.venue))
        .where(Shift.id == shift_id)
    )
    return result.scalar_one()


# ─── User / Profile ──────────────────────────────────────────────────────────

@router.get("/me", response_model=UserOut)
async def get_me(user: User = Depends(get_current_user)):
    return user


@router.get("/venues/active", response_model=list[VenueOut])
async def list_active_venues(
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    result = await session.execute(
        select(Venue).where(Venue.is_active == True).order_by(Venue.name)
    )
    return result.scalars().all()


# ─── Shifts ──────────────────────────────────────────────────────────────────

@router.get("/venues/stats", response_model=list[VenueStatsRow])
async def list_venue_stats(
    month: int | None = Query(default=None, ge=1, le=12),
    year: int | None = Query(default=None, ge=2000, le=2100),
    include_inactive: bool = Query(default=False),
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if not (
        _can_manage_all_venue_shifts(user)
        or has_permission(user, "can_manage_team")
        or has_permission(user, "can_view_team_payroll")
    ):
        raise HTTPException(status_code=403, detail="Недостаточно прав для просмотра статистики точек")

    now = datetime.now(timezone.utc)
    selected_month = month or now.month
    selected_year = year or now.year

    venues_query = select(Venue)
    if not include_inactive:
        venues_query = venues_query.where(Venue.is_active == True)
    if not _can_manage_all_venue_shifts(user):
        venues_query = venues_query.where(Venue.id == user.venue_id)
    venues_result = await session.execute(venues_query.order_by(Venue.name, Venue.id))
    venues = venues_result.scalars().all()
    venue_ids = [venue.id for venue in venues]
    if not venue_ids:
        return []

    assigned_result = await session.execute(
        select(User.venue_id, func.count(User.id))
        .where(User.is_active == True, User.venue_id.in_(venue_ids))
        .group_by(User.venue_id)
    )
    assigned_by_venue = {venue_id: int(count or 0) for venue_id, count in assigned_result.all()}

    approved_status = Shift.status == "approved"
    pending_status = Shift.status == "pending"
    shift_result = await session.execute(
        select(
            Shift.venue_id,
            func.count(func.distinct(case((approved_status, Shift.user_id), else_=None))),
            func.count(case((approved_status, Shift.id), else_=None)),
            func.count(case((pending_status, Shift.id), else_=None)),
            func.coalesce(func.sum(case((approved_status, Shift.total_hours), else_=0)), 0),
            func.coalesce(func.sum(case((approved_status, Shift.salary_earned), else_=0)), 0),
        )
        .where(
            Shift.venue_id.in_(venue_ids),
            func.extract("month", Shift.date) == selected_month,
            func.extract("year", Shift.date) == selected_year,
        )
        .group_by(Shift.venue_id)
    )
    shifts_by_venue = {
        venue_id: {
            "worked": int(worked or 0),
            "approved": int(approved or 0),
            "pending": int(pending or 0),
            "hours": safe_decimal(hours),
            "accruals": safe_decimal(accruals),
        }
        for venue_id, worked, approved, pending, hours, accruals in shift_result.all()
    }

    adjustment_result = await session.execute(
        select(
            Adjustment.venue_id,
            func.coalesce(
                func.sum(case((Adjustment.type == AdjustmentType.bonus, Adjustment.amount), else_=0)),
                0,
            ),
            func.coalesce(
                func.sum(case((Adjustment.type == AdjustmentType.penalty, Adjustment.amount), else_=0)),
                0,
            ),
        )
        .where(
            Adjustment.venue_id.in_(venue_ids),
            Adjustment.month == selected_month,
            Adjustment.year == selected_year,
        )
        .group_by(Adjustment.venue_id)
    )
    adjustments_by_venue = {
        venue_id: {
            "bonuses": safe_decimal(bonuses),
            "deductions": safe_decimal(deductions),
        }
        for venue_id, bonuses, deductions in adjustment_result.all()
    }

    rows: list[VenueStatsRow] = []
    for venue in venues:
        shift_stats = shifts_by_venue.get(venue.id, {})
        adjustment_stats = adjustments_by_venue.get(venue.id, {})
        shift_accruals = shift_stats.get("accruals", Decimal("0.00"))
        bonuses = adjustment_stats.get("bonuses", Decimal("0.00"))
        deductions = adjustment_stats.get("deductions", Decimal("0.00"))
        total_accruals = calculate_payout_total(shift_accruals, bonuses, deductions)

        rows.append(
            VenueStatsRow(
                venue_id=venue.id,
                venue_name=venue.name,
                is_active=venue.is_active,
                assigned_employees_count=assigned_by_venue.get(venue.id, 0),
                worked_employees_count=shift_stats.get("worked", 0),
                approved_shifts_count=shift_stats.get("approved", 0),
                pending_shifts_count=shift_stats.get("pending", 0),
                approved_hours=shift_stats.get("hours", Decimal("0.00")),
                shift_accruals=shift_accruals,
                bonuses=bonuses,
                deductions=deductions,
                total_accruals=total_accruals,
            )
        )
    return rows


@router.post("/shifts", response_model=ShiftOut)
async def create_shift(
    shift_data: ShiftCreate,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    today = date.today()
    yesterday = today - timedelta(days=1)

    # Baristas and cooks can only create shifts for today or yesterday
    if user.role in (UserRole.barista, UserRole.cook) and shift_data.date < yesterday:
        raise HTTPException(
            status_code=403,
            detail="You can only create shifts for today or yesterday",
        )

    existing_shift_result = await session.execute(
        select(Shift.id).where(
            Shift.user_id == user.id,
            Shift.date == shift_data.date,
            Shift.status.in_(("pending", "approved")),
        )
    )
    existing_shift_id = existing_shift_result.scalar_one_or_none()
    if existing_shift_id:
        raise HTTPException(
            status_code=409,
            detail="Смена за этот день уже создана. Дождитесь подтверждения или обратитесь к администратору.",
        )

    actual_venue_id = shift_data.venue_id or user.venue_id
    await _get_active_venue(session, actual_venue_id)

    # Calculate hours and salary
    total_hours = calculate_hours(shift_data.start_time, shift_data.end_time)
    salary_earned = calculate_salary(
        total_hours, user.hourly_rate,
        revenue=shift_data.revenue,
        revenue_percentage=user.revenue_percentage,
        pay_model=user.pay_model.value,
    )

    shift = Shift(
        user_id=user.id,
        venue_id=actual_venue_id,
        date=shift_data.date,
        start_time=shift_data.start_time,
        end_time=shift_data.end_time,
        cashier_hours=shift_data.cashier_hours,
        total_hours=total_hours,
        salary_earned=salary_earned,
        revenue=shift_data.revenue,
        comment=shift_data.comment,
    )
    session.add(shift)
    await session.commit()
    shift = await _load_shift_response(session, shift.id)

    # Audit logging should not break a successfully saved shift.
    try:
        log = AuditLog(
            user_id=user.id,
            venue_id=shift.venue_id,
            action="shift_created",
            entity_type="shift",
            entity_id=shift.id,
            new_value={
                "date": str(shift.date),
                "start_time": str(shift.start_time),
                "end_time": str(shift.end_time),
                "salary": str(shift.salary_earned),
                "venue_id": str(shift.venue_id),
            },
        )
        session.add(log)
        await session.commit()
    except Exception:
        await session.rollback()
        logger.exception("Audit log write failed after successful shift creation", extra={"shift_id": str(shift.id)})

    return shift


@router.get("/shifts", response_model=list[ShiftOut])
async def list_shifts(
    month: int | None = None,
    year: int | None = None,
    venue_id: uuid.UUID | None = Query(default=None),
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    now = datetime.now(timezone.utc)
    m = month or now.month
    y = year or now.year

    query = (
        select(Shift)
        .outerjoin(User, Shift.user_id == User.id)
        .options(selectinload(Shift.user), selectinload(Shift.venue))
        .where(
            func.extract("month", Shift.date) == m,
            func.extract("year", Shift.date) == y,
        )
    )

    if venue_id is not None and _can_manage_all_venue_shifts(user):
        query = query.where(Shift.venue_id == venue_id)
    elif _can_manage_all_venue_shifts(user):
        pass
    elif _can_view_team_history_scope(user):
        query = query.where(Shift.venue_id == user.venue_id)
    else:
        query = query.where(Shift.user_id == user.id)

    query = query.order_by(Shift.date.desc(), Shift.start_time.desc())

    result = await session.execute(query)
    shifts = result.scalars().all()
    return shifts


@router.get("/shifts/pending", response_model=list[ShiftOut])
async def list_pending_shifts(
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """Admin/senior: list all pending shifts for the venue."""
    if not (has_permission(user, "can_approve_shifts") or has_permission(user, "can_edit_team_shifts")):
        raise HTTPException(status_code=403, detail="Only users with shift approval rights can view pending shifts")

    query = (
        select(Shift)
        .options(selectinload(Shift.user), selectinload(Shift.venue))
        .where(Shift.status == "pending")
    )

    if not _can_manage_all_venue_shifts(user):
        query = query.where(Shift.venue_id == user.venue_id)

    query = query.order_by(Shift.date.desc(), Shift.start_time.desc())

    result = await session.execute(query)
    shifts = result.scalars().all()
    return shifts


@router.patch("/shifts/{shift_id}", response_model=ShiftOut)
async def update_shift(
    shift_id: uuid.UUID,
    shift_data: ShiftUpdate,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    query = (
        select(Shift)
        .options(selectinload(Shift.user), selectinload(Shift.venue))
        .where(Shift.id == shift_id)
    )
    if not _can_manage_all_venue_shifts(user):
        query = query.where(Shift.venue_id == user.venue_id)

    result = await session.execute(query)
    shift = result.scalar_one_or_none()
    if not shift:
        raise HTTPException(status_code=404, detail="Shift not found")

    # Only admin/senior can approve/update shifts
    if not (has_permission(user, "can_approve_shifts") or has_permission(user, "can_edit_team_shifts")):
        raise HTTPException(status_code=403, detail="Only users with shift edit rights can update shifts")

    old_status = shift.status
    old_venue_id = shift.venue_id

    if shift_data.venue_id is not None and shift_data.venue_id != shift.venue_id:
        if shift.status != "pending":
            raise HTTPException(status_code=409, detail="Точку можно изменить только у смены на подтверждении")
        await _get_active_venue(session, shift_data.venue_id)
        shift.venue_id = shift_data.venue_id

    if shift_data.start_time is not None:
        shift.start_time = shift_data.start_time
    if shift_data.end_time is not None:
        shift.end_time = shift_data.end_time
    if shift_data.cashier_hours is not None:
        shift.cashier_hours = shift_data.cashier_hours
    if shift_data.revenue is not None:
        shift.revenue = shift_data.revenue
    if shift_data.comment is not None:
        shift.comment = shift_data.comment
    if shift_data.status is not None:
        if shift_data.status not in ("pending", "approved", "rejected"):
            raise HTTPException(status_code=400, detail=f"Invalid status: {shift_data.status}")
        old_status = shift.status
        shift.status = shift_data.status

    # Recalculate if times or revenue changed
    if shift_data.start_time is not None or shift_data.end_time is not None or shift_data.revenue is not None:
        shift.total_hours = calculate_hours(shift.start_time, shift.end_time)
        # Get user's pay model
        user_result = await session.get(User, shift.user_id)
        if user_result:
            shift.salary_earned = calculate_salary(
                shift.total_hours, user_result.hourly_rate,
                revenue=shift.revenue or shift_data.revenue,
                revenue_percentage=user_result.revenue_percentage,
                pay_model=user_result.pay_model.value,
            )

    await session.commit()
    shift = await _load_shift_response(session, shift.id)

    # Audit log
    action = "shift_edited"
    if shift_data.status == "approved":
        action = "shift_approved"
    elif shift_data.status == "rejected":
        action = "shift_rejected"

    try:
        old_value = {"status": old_status} if shift_data.status else {}
        new_value = {"status": shift.status, "salary": str(shift.salary_earned)}
        if old_venue_id != shift.venue_id:
            old_value["venue_id"] = str(old_venue_id)
            new_value["venue_id"] = str(shift.venue_id)

        log = AuditLog(
            user_id=user.id,
            target_user_id=shift.user_id,
            venue_id=shift.venue_id,
            action=action,
            entity_type="shift",
            entity_id=shift.id,
            old_value=old_value or None,
            new_value=new_value,
        )
        session.add(log)
        await session.commit()
    except Exception:
        await session.rollback()
        logger.exception("Audit log write failed after successful shift update", extra={"shift_id": str(shift.id)})

    # Send notification to shift owner
    if shift_data.status and shift.user_id:
        try:
            shift_owner = await session.get(User, shift.user_id)
            if shift_owner and shift_owner.telegram_id:
                if shift_data.status == "approved":
                    await notify_shift_approved(
                        shift_owner.telegram_id,
                        str(shift.date),
                        str(shift.salary_earned),
                    )
                elif shift_data.status == "rejected":
                    await notify_shift_rejected(
                        shift_owner.telegram_id,
                        str(shift.date),
                        shift_data.comment or "",
                    )
        except Exception:
            logger.exception("Shift notification failed after successful shift update", extra={"shift_id": str(shift.id)})

    return shift


@router.get("/payroll/summary", response_model=PayrollSummaryOut)
async def payroll_summary(
    month: int | None = None,
    year: int | None = None,
    venue_id: uuid.UUID | None = Query(default=None),
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if not has_permission(user, "can_view_team_payroll"):
        raise HTTPException(status_code=403, detail="Only users with payroll access can view payroll summary")

    now = datetime.now(timezone.utc)
    m = month or now.month
    y = year or now.year

    shifts_query = (
        select(Shift, User)
        .outerjoin(User, Shift.user_id == User.id)
        .where(
            func.extract("month", Shift.date) == m,
            func.extract("year", Shift.date) == y,
        )
    )
    if venue_id is not None and _can_manage_all_venue_shifts(user):
        shifts_query = shifts_query.where(Shift.venue_id == venue_id)
    elif not _can_manage_all_venue_shifts(user):
        shifts_query = shifts_query.where(Shift.venue_id == user.venue_id)
    shifts_query = shifts_query.order_by(User.name, Shift.date)

    shifts_result = await session.execute(shifts_query)
    shifts_with_users = shifts_result.all()

    adjustments_query = (
        select(Adjustment, User)
        .join(User, Adjustment.user_id == User.id)
        .where(
            Adjustment.month == m,
            Adjustment.year == y,
        )
    )
    if venue_id is not None and _can_manage_all_venue_shifts(user):
        adjustments_query = adjustments_query.where(Adjustment.venue_id == venue_id)
    elif not _can_manage_all_venue_shifts(user):
        adjustments_query = adjustments_query.where(Adjustment.venue_id == user.venue_id)
    adjustments_query = adjustments_query.order_by(User.name)

    adjustments_result = await session.execute(adjustments_query)
    adjustments_with_users = adjustments_result.all()

    rows_by_user: dict[uuid.UUID, dict] = {}

    pending_shifts_count = 0
    approved_shifts_count = 0

    for shift, shift_user in shifts_with_users:
        if shift.status == "pending":
            pending_shifts_count += 1
            continue

        if shift.status != "approved":
            continue

        approved_shifts_count += 1
        row = rows_by_user.setdefault(
            shift.user_id,
            {
                "user_id": shift.user_id,
                "user_name": shift_user.name,
                "approved_shifts_count": 0,
                "total_hours": Decimal("0.00"),
                "shift_payout": Decimal("0.00"),
                "bonuses": Decimal("0.00"),
                "penalties": Decimal("0.00"),
            },
        )
        row["approved_shifts_count"] += 1
        row["total_hours"] += shift.total_hours
        row["shift_payout"] += shift.salary_earned

    for adjustment, adjustment_user in adjustments_with_users:
        row = rows_by_user.setdefault(
            adjustment.user_id,
            {
                "user_id": adjustment.user_id,
                "user_name": adjustment_user.name,
                "approved_shifts_count": 0,
                "total_hours": Decimal("0.00"),
                "shift_payout": Decimal("0.00"),
                "bonuses": Decimal("0.00"),
                "penalties": Decimal("0.00"),
            },
        )
        if adjustment.type == AdjustmentType.bonus:
            row["bonuses"] += adjustment.amount
        else:
            row["penalties"] += adjustment.amount

    rows: list[PayrollSummaryRow] = []
    total_hours = Decimal("0.00")
    total_shift_payout = Decimal("0.00")
    total_bonuses = Decimal("0.00")
    total_penalties = Decimal("0.00")

    for row in rows_by_user.values():
        total_payout = calculate_payout_total(
            row["shift_payout"], row["bonuses"], row["penalties"]
        )
        total_hours += row["total_hours"]
        total_shift_payout += row["shift_payout"]
        total_bonuses += row["bonuses"]
        total_penalties += row["penalties"]

        if (
            row["approved_shifts_count"] == 0
            and row["bonuses"] == Decimal("0.00")
            and row["penalties"] == Decimal("0.00")
        ):
            continue

        rows.append(
            PayrollSummaryRow(
                user_id=row["user_id"],
                user_name=row["user_name"],
                approved_shifts_count=row["approved_shifts_count"],
                total_hours=row["total_hours"],
                shift_payout=row["shift_payout"],
                bonuses=row["bonuses"],
                penalties=row["penalties"],
                total_payout=total_payout,
            )
        )

    rows.sort(key=lambda item: (-item.total_payout, item.user_name.lower()))

    employees_count = sum(1 for row in rows if row.approved_shifts_count > 0)

    return PayrollSummaryOut(
        month=m,
        year=y,
        employees_count=employees_count,
        pending_shifts_count=pending_shifts_count,
        approved_shifts_count=approved_shifts_count,
        total_hours=total_hours,
        total_shift_payout=total_shift_payout,
        total_bonuses=total_bonuses,
        total_penalties=total_penalties,
        total_payout=calculate_payout_total(
            total_shift_payout, total_bonuses, total_penalties
        ),
        rows=rows,
    )


# ─── Expenses ────────────────────────────────────────────────────────────────

@router.get("/payroll-runs/preview", response_model=PayrollPreviewOut)
async def payroll_run_preview(
    period_start: date,
    period_end: date,
    venue_id: uuid.UUID | None = Query(default=None),
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if not has_permission(user, "can_view_team_payroll"):
        raise HTTPException(status_code=403, detail="Only users with payroll access can preview payroll")
    if period_start > period_end:
        raise HTTPException(status_code=400, detail="period_start must be before or equal to period_end")

    shifts_query = (
        select(Shift, User)
        .outerjoin(User, Shift.user_id == User.id)
        .options(selectinload(Shift.venue))
        .where(
            Shift.status == "approved",
            Shift.date >= period_start,
            Shift.date <= period_end,
        )
    )
    if venue_id is not None and _can_manage_all_venue_shifts(user):
        shifts_query = shifts_query.where(Shift.venue_id == venue_id)
    elif not _can_manage_all_venue_shifts(user):
        shifts_query = shifts_query.where(Shift.venue_id == user.venue_id)
    shifts_result = await session.execute(shifts_query.order_by(User.name, Shift.date))
    approved_shifts = shifts_result.all()

    adjustments_query = (
        select(Adjustment, User)
        .join(User, Adjustment.user_id == User.id)
        .options(selectinload(Adjustment.venue))
        .where(
            Adjustment.year * 100 + Adjustment.month >= period_start.year * 100 + period_start.month,
            Adjustment.year * 100 + Adjustment.month <= period_end.year * 100 + period_end.month,
        )
    )
    if venue_id is not None and _can_manage_all_venue_shifts(user):
        adjustments_query = adjustments_query.where(Adjustment.venue_id == venue_id)
    elif not _can_manage_all_venue_shifts(user):
        adjustments_query = adjustments_query.where(Adjustment.venue_id == user.venue_id)
    adjustments_result = await session.execute(adjustments_query)
    adjustments = adjustments_result.all()

    rows_by_user: dict[uuid.UUID, dict] = {}
    for shift, shift_user in approved_shifts:
        if shift_user is None:
            continue
        row = rows_by_user.setdefault(
            shift.user_id,
            {
                "user_id": shift.user_id,
                "user_name": safe_text(shift_user.name, "Сотрудник"),
                "venue_names": set(),
                "shifts_count": 0,
                "total_hours": Decimal("0.00"),
                "base_amount": Decimal("0.00"),
                "bonuses": Decimal("0.00"),
                "deductions": Decimal("0.00"),
            },
        )
        shift_venue_name = safe_text(getattr(getattr(shift, "venue", None), "name", None), "")
        if shift_venue_name:
            row["venue_names"].add(shift_venue_name)
        row["shifts_count"] += 1
        row["total_hours"] += safe_decimal(shift.total_hours)
        row["base_amount"] += safe_decimal(shift.salary_earned)

    for adjustment, adjustment_user in adjustments:
        if adjustment_user is None:
            continue
        row = rows_by_user.setdefault(
            adjustment.user_id,
            {
                "user_id": adjustment.user_id,
                "user_name": safe_text(adjustment_user.name, "Сотрудник"),
                "venue_names": set(),
                "shifts_count": 0,
                "total_hours": Decimal("0.00"),
                "base_amount": Decimal("0.00"),
                "bonuses": Decimal("0.00"),
                "deductions": Decimal("0.00"),
            },
        )
        adjustment_venue_name = safe_text(getattr(getattr(adjustment, "venue", None), "name", None), "")
        if adjustment_venue_name:
            row["venue_names"].add(adjustment_venue_name)
        if adjustment.type == AdjustmentType.bonus:
            row["bonuses"] += safe_decimal(adjustment.amount)
        else:
            row["deductions"] += safe_decimal(adjustment.amount)

    preview_rows: list[PayrollPreviewRow] = []
    for row in rows_by_user.values():
        if not (row["shifts_count"] > 0 or row["bonuses"] or row["deductions"]):
            continue
        venue_names = sorted(row.pop("venue_names"))
        venue_name = (
            venue_names[0]
            if len(venue_names) == 1
            else "Несколько точек"
            if len(venue_names) > 1
            else "Точка не указана"
        )
        preview_rows.append(
            PayrollPreviewRow(
                **row,
                venue_name=venue_name,
                total_amount=calculate_payout_total(
                    row["base_amount"], row["bonuses"], row["deductions"]
                ),
            )
        )
    preview_rows.sort(key=lambda row: (row.user_name.lower(), row.user_id.hex))

    total_hours = sum((row.total_hours for row in preview_rows), Decimal("0.00"))
    total_base_amount = sum((row.base_amount for row in preview_rows), Decimal("0.00"))
    total_bonuses = sum((row.bonuses for row in preview_rows), Decimal("0.00"))
    total_deductions = sum((row.deductions for row in preview_rows), Decimal("0.00"))

    return PayrollPreviewOut(
        period_start=period_start,
        period_end=period_end,
        venue_id=venue_id,
        employees_count=len(preview_rows),
        shifts_count=sum(row.shifts_count for row in preview_rows),
        total_hours=total_hours,
        total_base_amount=total_base_amount,
        total_bonuses=total_bonuses,
        total_deductions=total_deductions,
        total_amount=calculate_payout_total(
            total_base_amount, total_bonuses, total_deductions
        ),
        rows=preview_rows,
    )


@router.post("/payroll-runs", response_model=PayrollRunRead, status_code=201)
async def create_payroll_run(
    payroll_data: PayrollRunCreate,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if user.role not in (UserRole.owner, UserRole.admin):
        raise HTTPException(status_code=403, detail="Only owners and admins can create payroll runs")
    if payroll_data.period_start > payroll_data.period_end:
        raise HTTPException(status_code=400, detail="period_start must be before or equal to period_end")
    if payroll_data.revenue_total is not None and payroll_data.venue_id is None:
        raise HTTPException(
            status_code=422,
            detail="Выручку можно указать только для расчёта конкретной точки",
        )

    revenue_total = (
        safe_decimal(payroll_data.revenue_total).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        if payroll_data.revenue_total is not None
        else None
    )

    source_shifts_query = (
        select(Shift.id)
        .where(
            Shift.status == "approved",
            Shift.date >= payroll_data.period_start,
            Shift.date <= payroll_data.period_end,
        )
    )
    if payroll_data.venue_id is not None:
        source_shifts_query = source_shifts_query.where(Shift.venue_id == payroll_data.venue_id)
    source_shifts_result = await session.execute(source_shifts_query)
    source_shift_ids = list(source_shifts_result.scalars().all())

    source_adjustments_query = select(Adjustment.id).where(
        Adjustment.year * 100 + Adjustment.month >= payroll_data.period_start.year * 100 + payroll_data.period_start.month,
        Adjustment.year * 100 + Adjustment.month <= payroll_data.period_end.year * 100 + payroll_data.period_end.month,
    )
    if payroll_data.venue_id is not None:
        source_adjustments_query = source_adjustments_query.where(
            Adjustment.venue_id == payroll_data.venue_id
        )
    source_adjustments_result = await session.execute(source_adjustments_query)
    source_adjustment_ids = list(source_adjustments_result.scalars().all())

    active_run_statuses = (
        PayrollRunStatus.draft,
        PayrollRunStatus.finalized,
        PayrollRunStatus.paid,
    )
    if source_shift_ids:
        shift_conflict_result = await session.execute(
            select(PayrollRunShiftSource.shift_id)
            .join(PayrollRun, PayrollRun.id == PayrollRunShiftSource.payroll_run_id)
            .where(
                PayrollRunShiftSource.shift_id.in_(source_shift_ids),
                PayrollRun.status.in_(active_run_statuses),
            )
            .limit(1)
        )
        if shift_conflict_result.scalar_one_or_none() is not None:
            raise HTTPException(status_code=409, detail="An approved shift is already included in another payroll run")
    if source_adjustment_ids:
        adjustment_conflict_result = await session.execute(
            select(PayrollRunAdjustmentSource.adjustment_id)
            .join(PayrollRun, PayrollRun.id == PayrollRunAdjustmentSource.payroll_run_id)
            .where(
                PayrollRunAdjustmentSource.adjustment_id.in_(source_adjustment_ids),
                PayrollRun.status.in_(active_run_statuses),
            )
            .limit(1)
        )
        if adjustment_conflict_result.scalar_one_or_none() is not None:
            raise HTTPException(status_code=409, detail="An adjustment is already included in another payroll run")

    preview = await payroll_run_preview(
        period_start=payroll_data.period_start,
        period_end=payroll_data.period_end,
        venue_id=payroll_data.venue_id,
        user=user,
        session=session,
    )
    title = payroll_data.title or (
        f"Payroll {payroll_data.period_start.isoformat()} - {payroll_data.period_end.isoformat()}"
    )
    payroll_run = PayrollRun(
        title=title,
        period_start=payroll_data.period_start,
        period_end=payroll_data.period_end,
        status=PayrollRunStatus.draft,
        total_amount=preview.total_amount,
        total_paid=Decimal("0.00"),
        revenue_total=revenue_total,
        created_by_id=user.id,
        venue_id=payroll_data.venue_id,
        notes=payroll_data.notes,
    )
    payroll_run.items = [
        PayrollRunItem(
            user_id=row.user_id,
            approved_shifts_count=row.shifts_count,
            approved_hours=row.total_hours,
            base_amount=row.base_amount,
            bonus_amount=row.bonuses,
            deduction_amount=row.deductions,
            final_amount=row.total_amount,
            paid_amount=Decimal("0.00"),
            remaining_amount=row.total_amount,
        )
        for row in preview.rows
    ]
    payroll_run.shift_sources = [
        PayrollRunShiftSource(shift_id=shift_id) for shift_id in source_shift_ids
    ]
    payroll_run.adjustment_sources = [
        PayrollRunAdjustmentSource(adjustment_id=adjustment_id)
        for adjustment_id in source_adjustment_ids
    ]

    try:
        session.add(payroll_run)
        await session.flush()
        await session.commit()
    except Exception:
        await session.rollback()
        raise

    return await get_payroll_run(
        payroll_run_id=payroll_run.id,
        user=user,
        session=session,
    )


@router.get("/payroll-runs", response_model=list[PayrollRunListItem])
async def list_payroll_runs(
    venue_id: uuid.UUID | None = Query(default=None),
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if not _can_view_payroll_runs(user):
        raise HTTPException(status_code=403, detail="Only users with payroll access can view payroll runs")

    query = (
        select(PayrollRun)
        .options(
            selectinload(PayrollRun.items),
            selectinload(PayrollRun.venue),
            selectinload(PayrollRun.created_by_user),
        )
        .order_by(PayrollRun.period_end.desc(), PayrollRun.created_at.desc())
    )
    if _can_manage_all_venue_shifts(user):
        if venue_id is not None:
            query = query.where(PayrollRun.venue_id == venue_id)
    else:
        query = query.where(PayrollRun.venue_id == user.venue_id)
        if venue_id is not None and venue_id != user.venue_id:
            return []

    result = await session.execute(query)
    runs = result.scalars().all()
    return [
        PayrollRunListItem(
            id=run.id,
            title=run.title,
            period_start=run.period_start,
            period_end=run.period_end,
            venue_id=run.venue_id,
            venue_name=safe_text(getattr(run.venue, "name", None), "Основная точка") if run.venue else "Все точки",
            status=run.status.value if hasattr(run.status, "value") else str(run.status),
            employees_count=len(run.items),
            total_amount=run.total_amount,
            total_paid=run.total_paid,
            revenue_total=run.revenue_total,
            payroll_share_percent=calculate_payroll_share(
                run.total_amount, run.revenue_total
            ),
            created_by_id=run.created_by_id,
            created_by_name=safe_text(getattr(run.created_by_user, "name", None), "Пользователь"),
            created_at=run.created_at,
        )
        for run in runs
    ]


@router.get("/me/payroll-runs", response_model=list[PersonalPayrollRunRead])
async def list_my_payroll_runs(
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """Return only finalized or paid payroll snapshots belonging to the current user."""
    result = await session.execute(
        select(PayrollRun)
        .join(PayrollRunItem, PayrollRunItem.payroll_run_id == PayrollRun.id)
        .options(
            selectinload(PayrollRun.items),
            selectinload(PayrollRun.payments),
            selectinload(PayrollRun.venue),
        )
        .where(
            PayrollRunItem.user_id == user.id,
            PayrollRun.status.in_((PayrollRunStatus.finalized, PayrollRunStatus.paid)),
        )
        .order_by(PayrollRun.period_end.desc(), PayrollRun.created_at.desc())
    )
    runs = result.unique().scalars().all()
    personal_runs: list[PersonalPayrollRunRead] = []
    for run in runs:
        item = next((run_item for run_item in run.items if run_item.user_id == user.id), None)
        if item is None:
            continue
        payments = [
            PersonalPayrollPaymentRead(
                amount=payment.amount,
                payment_date=payment.payment_date,
                method=payment.method,
                comment=payment.comment,
                created_at=payment.created_at,
            )
            for payment in run.payments
            if payment.user_id == user.id
        ]
        payments.sort(key=lambda payment: (payment.payment_date, payment.created_at), reverse=True)
        personal_runs.append(
            PersonalPayrollRunRead(
                payroll_run_id=run.id,
                title=run.title,
                period_start=run.period_start,
                period_end=run.period_end,
                venue_name=safe_text(getattr(run.venue, "name", None), "Основная точка") if run.venue else "Основная точка",
                status=run.status.value if hasattr(run.status, "value") else str(run.status),
                final_amount=item.final_amount,
                paid_amount=item.paid_amount,
                remaining_amount=item.remaining_amount,
                payments=payments,
            )
        )
    return personal_runs


@router.get("/payroll-runs/{payroll_run_id}", response_model=PayrollRunRead)
async def get_payroll_run(
    payroll_run_id: uuid.UUID,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if not _can_view_payroll_runs(user):
        raise HTTPException(status_code=403, detail="Only users with payroll access can view payroll runs")

    result = await session.execute(
        select(PayrollRun)
        .options(
            selectinload(PayrollRun.items).selectinload(PayrollRunItem.user),
            selectinload(PayrollRun.payments),
            selectinload(PayrollRun.venue),
            selectinload(PayrollRun.created_by_user),
        )
        .where(PayrollRun.id == payroll_run_id)
    )
    run = result.scalar_one_or_none()
    if run is None or not _payroll_run_is_visible(run, user):
        raise HTTPException(status_code=404, detail="Payroll run not found")

    return PayrollRunRead(
        id=run.id,
        title=run.title,
        period_start=run.period_start,
        period_end=run.period_end,
        status=run.status.value if hasattr(run.status, "value") else str(run.status),
        total_amount=run.total_amount,
        total_paid=run.total_paid,
        revenue_total=run.revenue_total,
        payroll_share_percent=calculate_payroll_share(
            run.total_amount, run.revenue_total
        ),
        created_by_id=run.created_by_id,
        venue_id=run.venue_id,
        venue_name=safe_text(getattr(run.venue, "name", None), "Основная точка") if run.venue else "Все точки",
        created_by_name=safe_text(getattr(run.created_by_user, "name", None), "Пользователь"),
        created_at=run.created_at,
        finalized_at=run.finalized_at,
        paid_at=run.paid_at,
        notes=run.notes,
        items=[
            PayrollRunItemRead(
                id=item.id,
                payroll_run_id=item.payroll_run_id,
                user_id=item.user_id,
                user_name=safe_text(getattr(item.user, "name", None), "Сотрудник"),
                approved_shifts_count=item.approved_shifts_count,
                approved_hours=item.approved_hours,
                base_amount=item.base_amount,
                bonus_amount=item.bonus_amount,
                deduction_amount=item.deduction_amount,
                final_amount=item.final_amount,
                paid_amount=item.paid_amount,
                remaining_amount=item.remaining_amount,
                created_at=item.created_at,
            )
            for item in run.items
        ],
        payments=[
            PayrollPaymentRead(
                id=payment.id,
                payroll_run_id=payment.payroll_run_id,
                user_id=payment.user_id,
                amount=payment.amount,
                payment_date=payment.payment_date,
                method=payment.method,
                comment=payment.comment,
                created_by_id=payment.created_by_id,
                created_at=payment.created_at,
            )
            for payment in run.payments
        ],
    )


@router.patch("/payroll-runs/{payroll_run_id}/revenue", response_model=PayrollRunRead)
async def update_payroll_run_revenue(
    payroll_run_id: uuid.UUID,
    revenue_data: PayrollRunRevenueUpdate,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if user.role not in (UserRole.owner, UserRole.admin):
        raise HTTPException(status_code=403, detail="Только владелец или администратор может изменить выручку")

    result = await session.execute(
        select(PayrollRun).where(PayrollRun.id == payroll_run_id)
    )
    run = result.scalar_one_or_none()
    if run is None or not _payroll_run_is_visible(run, user):
        raise HTTPException(status_code=404, detail="Расчёт не найден")
    if run.venue_id is None:
        raise HTTPException(
            status_code=422,
            detail="Выручку можно указать только для расчёта конкретной точки",
        )
    if run.status != PayrollRunStatus.draft:
        raise HTTPException(
            status_code=409,
            detail="Выручку можно изменить только в черновике расчёта",
        )

    previous_revenue = run.revenue_total
    run.revenue_total = (
        safe_decimal(revenue_data.revenue_total).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        if revenue_data.revenue_total is not None
        else None
    )
    try:
        await session.commit()
    except Exception:
        await session.rollback()
        raise

    payroll_share_percent = calculate_payroll_share(
        run.total_amount, run.revenue_total
    )
    try:
        session.add(
            AuditLog(
                user_id=user.id,
                venue_id=run.venue_id,
                action="payroll_revenue_updated",
                entity_type="payroll_run",
                entity_id=run.id,
                old_value={
                    "revenue_total": str(previous_revenue) if previous_revenue is not None else None,
                },
                new_value={
                    "revenue_total": str(run.revenue_total) if run.revenue_total is not None else None,
                    "payroll_share_percent": str(payroll_share_percent) if payroll_share_percent is not None else None,
                },
            )
        )
        await session.commit()
    except Exception:
        await session.rollback()
        logger.exception(
            "Audit log write failed after payroll revenue update",
            extra={"payroll_run_id": str(run.id)},
        )

    return await get_payroll_run(
        payroll_run_id=run.id,
        user=user,
        session=session,
    )


@router.post("/payroll-runs/{payroll_run_id}/finalize", response_model=PayrollRunRead)
async def finalize_payroll_run(
    payroll_run_id: uuid.UUID,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if user.role not in (UserRole.owner, UserRole.admin):
        raise HTTPException(status_code=403, detail="Only owners and admins can finalize payroll runs")

    result = await session.execute(
        select(PayrollRun)
        .options(selectinload(PayrollRun.payments))
        .where(PayrollRun.id == payroll_run_id)
    )
    run = result.scalar_one_or_none()
    if run is None or not _payroll_run_is_visible(run, user):
        raise HTTPException(status_code=404, detail="Payroll run not found")
    if run.status != PayrollRunStatus.draft:
        raise HTTPException(status_code=409, detail="Only draft payroll runs can be finalized")

    run.status = PayrollRunStatus.finalized
    run.finalized_at = datetime.now(timezone.utc)
    try:
        await session.commit()
    except Exception:
        await session.rollback()
        raise

    return await get_payroll_run(payroll_run_id=payroll_run_id, user=user, session=session)


@router.post("/payroll-runs/{payroll_run_id}/cancel", response_model=PayrollRunRead)
async def cancel_payroll_run(
    payroll_run_id: uuid.UUID,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if user.role not in (UserRole.owner, UserRole.admin):
        raise HTTPException(status_code=403, detail="Only owners and admins can cancel payroll runs")

    result = await session.execute(
        select(PayrollRun)
        .options(selectinload(PayrollRun.payments))
        .where(PayrollRun.id == payroll_run_id)
    )
    run = result.scalar_one_or_none()
    if run is None or not _payroll_run_is_visible(run, user):
        raise HTTPException(status_code=404, detail="Payroll run not found")
    if run.status == PayrollRunStatus.paid:
        raise HTTPException(status_code=409, detail="Paid payroll runs cannot be changed")
    if run.status == PayrollRunStatus.cancelled:
        raise HTTPException(status_code=409, detail="Payroll run is already cancelled")
    if run.status == PayrollRunStatus.finalized and run.payments:
        raise HTTPException(status_code=409, detail="Finalized payroll runs with payments cannot be cancelled")

    run.status = PayrollRunStatus.cancelled
    try:
        await session.commit()
    except Exception:
        await session.rollback()
        raise

    return await get_payroll_run(payroll_run_id=payroll_run_id, user=user, session=session)


@router.post("/payroll-runs/{payroll_run_id}/payments", response_model=PayrollPaymentResult, status_code=201)
async def create_payroll_payment(
    payroll_run_id: uuid.UUID,
    payment_data: PayrollPaymentCreate,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if user.role not in (UserRole.owner, UserRole.admin):
        raise HTTPException(status_code=403, detail="Only owners and admins can record payroll payments")

    try:
        run_result = await session.execute(
            select(PayrollRun)
            .with_for_update()
            .where(PayrollRun.id == payroll_run_id)
        )
        run = run_result.scalar_one_or_none()
        if run is None or not _payroll_run_is_visible(run, user):
            raise HTTPException(status_code=404, detail="Payroll run not found")
        if run.status != PayrollRunStatus.finalized:
            raise HTTPException(status_code=409, detail="Payments can only be recorded for finalized payroll runs")

        item_result = await session.execute(
            select(PayrollRunItem)
            .with_for_update()
            .where(
                PayrollRunItem.payroll_run_id == payroll_run_id,
                PayrollRunItem.user_id == payment_data.user_id,
            )
        )
        item = item_result.scalar_one_or_none()
        if item is None:
            raise HTTPException(status_code=404, detail="Employee is not included in this payroll run")

        amount = safe_decimal(payment_data.amount).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        current_paid = safe_decimal(item.paid_amount)
        current_remaining = safe_decimal(item.remaining_amount)
        if current_paid < 0 or current_remaining < 0:
            raise HTTPException(status_code=409, detail="Payroll item has invalid payment amounts")
        if amount <= 0:
            raise HTTPException(status_code=422, detail="Payment amount must be greater than zero")
        if amount > current_remaining:
            raise HTTPException(status_code=409, detail="Payment amount exceeds the employee remaining amount")

        new_paid = current_paid + amount
        new_remaining = current_remaining - amount
        new_total_paid = safe_decimal(run.total_paid) + amount
        if new_total_paid > safe_decimal(run.total_amount):
            raise HTTPException(status_code=409, detail="Payment amount exceeds the payroll run total")

        item.paid_amount = new_paid
        item.remaining_amount = new_remaining
        run.total_paid = new_total_paid

        all_items_result = await session.execute(
            select(PayrollRunItem)
            .with_for_update()
            .where(PayrollRunItem.payroll_run_id == payroll_run_id)
        )
        all_items = all_items_result.scalars().all()
        if all_items and all(safe_decimal(row.remaining_amount) == Decimal("0.00") for row in all_items):
            run.status = PayrollRunStatus.paid
            run.paid_at = datetime.now(timezone.utc)

        payment = PayrollPayment(
            payroll_run_id=payroll_run_id,
            user_id=payment_data.user_id,
            amount=amount,
            payment_date=payment_data.payment_date,
            method=payment_data.method,
            comment=payment_data.comment,
            created_by_id=user.id,
        )
        session.add(payment)
        await session.flush()
        await session.commit()
    except HTTPException:
        await session.rollback()
        raise
    except Exception:
        await session.rollback()
        raise

    return PayrollPaymentResult(
        payment=PayrollPaymentRead(
            id=payment.id,
            payroll_run_id=payment.payroll_run_id,
            user_id=payment.user_id,
            amount=payment.amount,
            payment_date=payment.payment_date,
            method=payment.method,
            comment=payment.comment,
            created_by_id=payment.created_by_id,
            created_at=payment.created_at,
        ),
        user_id=item.user_id,
        paid_amount=item.paid_amount,
        remaining_amount=item.remaining_amount,
        total_paid=run.total_paid,
        status=run.status.value if hasattr(run.status, "value") else str(run.status),
    )


@router.post("/expenses", response_model=ExpenseOut)
async def create_expense(
    expense_data: ExpenseCreate,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    expense = Expense(
        user_id=user.id,
        venue_id=user.venue_id,
        amount=expense_data.amount,
        category=expense_data.category,
        comment=expense_data.comment,
        date=expense_data.date,
    )
    session.add(expense)
    await session.commit()
    await session.refresh(expense)
    return expense


@router.get("/expenses", response_model=list[ExpenseOut])
async def list_expenses(
    month: int | None = None,
    year: int | None = None,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    now = datetime.now(timezone.utc)
    m = month or now.month
    y = year or now.year

    query = select(Expense).where(
        Expense.user_id == user.id,
        func.extract("month", Expense.date) == m,
        func.extract("year", Expense.date) == y,
    ).order_by(Expense.date.desc())

    result = await session.execute(query)
    expenses = result.scalars().all()
    return expenses


# ─── Stats ───────────────────────────────────────────────────────────────────

@router.get("/stats/monthly", response_model=MonthlyStats)
async def monthly_stats(
    month: int | None = None,
    year: int | None = None,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    now = datetime.now(timezone.utc)
    m = month or now.month
    y = year or now.year

    # Shifts
    shifts_query = select(
        func.coalesce(func.sum(Shift.salary_earned), 0),
        func.coalesce(func.sum(Shift.total_hours), 0),
        func.coalesce(func.sum(Shift.cashier_hours), 0),
        func.count(Shift.id),
    ).where(
        Shift.user_id == user.id,
        Shift.status == "approved",
        func.extract("month", Shift.date) == m,
        func.extract("year", Shift.date) == y,
    )
    shift_result = await session.execute(shifts_query)
    total_earned, total_hours, total_cashier_hours, shifts_count = shift_result.one()

    # Expenses are operational records and are intentionally not payroll deductions.
    expenses_query = select(
        func.coalesce(func.sum(Expense.amount), 0),
    ).where(
        Expense.user_id == user.id,
        func.extract("month", Expense.date) == m,
        func.extract("year", Expense.date) == y,
    )
    expense_result = await session.execute(expenses_query)
    total_expenses = expense_result.scalar() or Decimal("0.00")

    # Bonuses
    bonuses_query = select(
        func.coalesce(func.sum(Adjustment.amount), 0),
    ).where(
        Adjustment.user_id == user.id,
        Adjustment.type == AdjustmentType.bonus,
        Adjustment.month == m,
        Adjustment.year == y,
    )
    bonuses_result = await session.execute(bonuses_query)
    total_bonuses = bonuses_result.scalar() or Decimal("0.00")

    # Penalties
    penalties_query = select(
        func.coalesce(func.sum(Adjustment.amount), 0),
    ).where(
        Adjustment.user_id == user.id,
        Adjustment.type == AdjustmentType.penalty,
        Adjustment.month == m,
        Adjustment.year == y,
    )
    penalties_result = await session.execute(penalties_query)
    total_penalties = penalties_result.scalar() or Decimal("0.00")
    total_payout = calculate_payout_total(total_earned, total_bonuses, total_penalties)

    return MonthlyStats(
        total_earned=Decimal(str(total_earned)),
        total_hours=Decimal(str(total_hours)),
        total_cashier_hours=Decimal(str(total_cashier_hours)),
        total_expenses=Decimal(str(total_expenses)),
        total_bonuses=Decimal(str(total_bonuses)),
        total_penalties=Decimal(str(total_penalties)),
        total_payout=total_payout,
        shifts_count=int(shifts_count),
    )


# ─── Audit Logs ─────────────────────────────────────────────────────────────

@router.get("/audit-logs", response_model=list[AuditLogOut])
async def list_audit_logs(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if not _can_view_general_audit(user):
        raise HTTPException(
            status_code=403,
            detail="Общий журнал действий доступен только пользователям с правом управления командой.",
        )

    offset = (page - 1) * limit

    query = (
        select(AuditLog)
        .options(selectinload(AuditLog.user), selectinload(AuditLog.target_user))
        .where(AuditLog.venue_id == user.venue_id)
        .order_by(AuditLog.created_at.desc())
        .offset(offset)
        .limit(limit)
    )
    result = await session.execute(query)
    logs = result.scalars().all()
    return _serialize_audit_logs(logs)


@router.get("/me/audit-log", response_model=list[AuditLogOut])
async def list_my_audit_log(
    limit: int = Query(20, ge=1, le=50),
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    query = (
        select(AuditLog)
        .options(selectinload(AuditLog.user), selectinload(AuditLog.target_user))
        .where(
            AuditLog.target_user_id == user.id,
            AuditLog.entity_type.in_(("user", "shift")),
        )
        .order_by(AuditLog.created_at.desc())
        .limit(limit)
    )
    result = await session.execute(query)
    logs = result.scalars().all()
    return _serialize_audit_logs(logs)


# ─── Adjustments (Bonuses / Penalties) ──────────────────────────────────────

@router.post("/adjustments", response_model=AdjustmentOut)
async def create_adjustment(
    adjustment_data: AdjustmentCreate,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    if not has_permission(user, "can_manage_adjustments"):
        raise HTTPException(status_code=403, detail="Only users with adjustment access can create adjustments")

    adjustment_venue_id = adjustment_data.venue_id or user.venue_id
    if not _can_manage_all_venue_shifts(user) and adjustment_venue_id != user.venue_id:
        raise HTTPException(status_code=403, detail="Корректировку можно отнести только к своей точке")
    adjustment_venue = await _get_active_venue(session, adjustment_venue_id)

    now = datetime.now(timezone.utc)
    adjustment = Adjustment(
        user_id=adjustment_data.user_id,
        venue_id=adjustment_venue_id,
        type=AdjustmentType(adjustment_data.type),
        amount=adjustment_data.amount,
        reason=adjustment_data.reason,
        created_by=user.id,
        month=now.month,
        year=now.year,
    )
    session.add(adjustment)
    await session.commit()
    await session.refresh(adjustment)

    # Audit log
    log = AuditLog(
        user_id=user.id,
        target_user_id=adjustment_data.user_id,
        venue_id=adjustment.venue_id,
        action=f"{adjustment_data.type}_added",
        entity_type="adjustment",
        entity_id=adjustment.id,
        new_value={
            "type": adjustment_data.type,
            "amount": str(adjustment_data.amount),
            "reason": adjustment_data.reason,
            "venue_id": str(adjustment.venue_id),
        },
    )
    session.add(log)
    await session.commit()

    # Send notification to target user
    target_user = await session.get(User, adjustment_data.user_id)
    if target_user and target_user.telegram_id:
        if adjustment_data.type == "bonus":
            await notify_bonus_added(
                target_user.telegram_id,
                str(adjustment_data.amount),
                adjustment_data.reason,
            )
        elif adjustment_data.type == "penalty":
            await notify_penalty_added(
                target_user.telegram_id,
                str(adjustment_data.amount),
                adjustment_data.reason,
            )

    return AdjustmentOut(
        id=adjustment.id,
        user_id=adjustment.user_id,
        venue_id=adjustment.venue_id,
        venue_name=adjustment_venue.name,
        type=adjustment.type.value,
        amount=adjustment.amount,
        reason=adjustment.reason,
        created_by=adjustment.created_by,
        month=adjustment.month,
        year=adjustment.year,
        created_at=adjustment.created_at.isoformat() if adjustment.created_at else "",
    )


@router.get("/adjustments", response_model=list[AdjustmentOut])
async def list_adjustments(
    month: int | None = None,
    year: int | None = None,
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    now = datetime.now(timezone.utc)
    m = month or now.month
    y = year or now.year

    query = (
        select(Adjustment)
        .options(
            selectinload(Adjustment.user),
            selectinload(Adjustment.creator),
            selectinload(Adjustment.venue),
        )
        .where(
            Adjustment.user_id == user.id,
            Adjustment.month == m,
            Adjustment.year == y,
        )
        .order_by(Adjustment.created_at.desc())
    )
    result = await session.execute(query)
    adjustments = result.scalars().all()

    return [
        AdjustmentOut(
            id=a.id,
            user_id=a.user_id,
            venue_id=a.venue_id,
            venue_name=a.venue.name if a.venue else None,
            type=a.type.value,
            amount=a.amount,
            reason=a.reason,
            created_by=a.created_by,
            month=a.month,
            year=a.year,
            created_at=a.created_at.isoformat() if a.created_at else "",
            user_name=a.user.name if a.user else None,
            creator_name=a.creator.name if a.creator else None,
        )
        for a in adjustments
    ]


# ─── CSV Export ──────────────────────────────────────────────────────────────

@router.get("/export/xlsx")
async def export_csv(
    month: int | None = None,
    year: int | None = None,
    venue_id: uuid.UUID | None = Query(default=None),
    user: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
):
    """Export monthly payroll data as CSV for accounting."""
    if not has_permission(user, "can_export_payroll"):
        raise HTTPException(status_code=403, detail="Only users with payroll export access can export payroll data")

    now = datetime.now(timezone.utc)
    m = month or now.month
    y = year or now.year

    shifts_query = (
        select(Shift, User)
        .outerjoin(User, Shift.user_id == User.id)
        .options(
            selectinload(Shift.venue),
        )
        .where(
            Shift.status == "approved",
            func.extract("month", Shift.date) == m,
            func.extract("year", Shift.date) == y,
        )
    )
    if venue_id is not None and _can_manage_all_venue_shifts(user):
        shifts_query = shifts_query.where(Shift.venue_id == venue_id)
    elif not _can_manage_all_venue_shifts(user):
        shifts_query = shifts_query.where(Shift.venue_id == user.venue_id)
    shifts_query = shifts_query.order_by(User.name, Shift.date)
    result = await session.execute(shifts_query)
    shifts_with_users = result.all()

    adj_query = (
        select(Adjustment, User)
        .join(User, Adjustment.user_id == User.id)
        .where(
            Adjustment.month == m,
            Adjustment.year == y,
        )
    )
    if venue_id is not None and _can_manage_all_venue_shifts(user):
        adj_query = adj_query.where(Adjustment.venue_id == venue_id)
    elif not _can_manage_all_venue_shifts(user):
        adj_query = adj_query.where(Adjustment.venue_id == user.venue_id)
    adj_result = await session.execute(adj_query)
    adjustments_with_users = adj_result.all()

    adj_by_user: dict[uuid.UUID, dict] = {}
    for adj, adj_user in adjustments_with_users:
        uid = adj.user_id
        if uid not in adj_by_user:
            adj_by_user[uid] = {"bonuses": Decimal("0"), "penalties": Decimal("0")}
        if adj.type == AdjustmentType.bonus:
            adj_by_user[uid]["bonuses"] += adj.amount
        else:
            adj_by_user[uid]["penalties"] += adj.amount

    month_names = [
        "январь",
        "февраль",
        "март",
        "апрель",
        "май",
        "июнь",
        "июль",
        "август",
        "сентябрь",
        "октябрь",
        "ноябрь",
        "декабрь",
    ]
    pay_model_labels = {
        "hourly": "Почасовая",
        "fixed_shift": "Фикс за смену",
        "revenue": "Процент от выручки",
        "hybrid": "Почасовая + процент",
    }
    status_labels = {
        "pending": "На подтверждении",
        "approved": "Утверждена",
        "rejected": "Отклонена",
    }

    period_text = f"{month_names[m - 1].capitalize()} {y}"
    if venue_id is not None and _can_manage_all_venue_shifts(user):
        venue_row = await session.get(Venue, venue_id)
        venue_text = safe_text(getattr(venue_row, "name", None), "Основная точка")
    elif not _can_manage_all_venue_shifts(user):
        venue_text = safe_text(getattr(getattr(user, "venue", None), "name", None), "Основная точка")
    else:
        venue_text = "Все точки"

    def pay_model_label(value) -> str:
        return pay_model_labels.get(normalize_pay_model(value), "Почасовая")

    def status_label(value) -> str:
        return status_labels.get(safe_text(value, "pending"), "Неизвестно")

    def add_title(sheet, columns: int) -> None:
        last_col = get_column_letter(columns)
        sheet.merge_cells(f"A1:{last_col}1")
        sheet.merge_cells(f"A2:{last_col}2")
        sheet["A1"] = "Порядок.Смены — отчёт по выплатам"
        sheet["A2"] = f"Период: {period_text} | Точка: {venue_text}"
        sheet["A1"].font = Font(bold=True, size=14)
        sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
        sheet["A2"].font = Font(size=11)
        sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")

    def apply_header_row(sheet, row_number: int, headers: list[str]) -> None:
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F6BFF", end_color="2F6BFF", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row_number, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    def set_formats(sheet, row_number: int, mapping: dict[int, str]) -> None:
        for col, fmt in mapping.items():
            sheet.cell(row=row_number, column=col).number_format = fmt

    def autofit(sheet, start_row: int, widths: dict[int, int]) -> None:
        for col_idx, base_width in widths.items():
            max_len = base_width
            for row_cells in sheet.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
                value = row_cells[0].value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 38)

    shift_rows: list[dict[str, object]] = []
    user_totals: dict[uuid.UUID, dict[str, object]] = {}

    for shift, shift_user in shifts_with_users:
        user_name = safe_text(getattr(shift_user, "name", None), "Сотрудник")
        position = safe_text(getattr(shift_user, "position", None), "")
        venue_name = "Точка не указана"
        pay_model_value = normalize_pay_model(getattr(shift_user, "pay_model", None) if shift_user is not None else None)

        shift_venue = getattr(shift, "venue", None)
        if shift_venue is not None:
            venue_name = safe_text(getattr(shift_venue, "name", None), venue_name)

        total_hours = safe_decimal(getattr(shift, "total_hours", None))
        hourly_rate = safe_decimal(getattr(shift_user, "hourly_rate", None) if shift_user is not None else None)
        revenue = safe_decimal(getattr(shift, "revenue", None))
        rev_pct = safe_decimal(getattr(shift_user, "revenue_percentage", None) if shift_user is not None else None)
        payout = calculate_salary(
            total_hours,
            hourly_rate,
            revenue=revenue if revenue != Decimal("0.00") else None,
            revenue_percentage=rev_pct if rev_pct != Decimal("0.00") else None,
            pay_model=pay_model_value,
        )
        user_id = shift.user_id
        user_adj = adj_by_user.get(user_id, {"bonuses": Decimal("0"), "penalties": Decimal("0")})

        if user_id not in user_totals:
            user_totals[user_id] = {
                "name": user_name,
                "shifts_count": 0,
                "hours": Decimal("0.00"),
                "shift_pay": Decimal("0.00"),
                "bonuses": user_adj["bonuses"],
                "penalties": user_adj["penalties"],
            }
        user_totals[user_id]["shifts_count"] += 1
        user_totals[user_id]["hours"] += total_hours
        user_totals[user_id]["shift_pay"] += payout

        shift_rows.append(
            {
                "employee": user_name,
                "venue": venue_name,
                "position": position,
                "date": getattr(shift, "date", None),
                "start_time": getattr(shift, "start_time", None),
                "end_time": getattr(shift, "end_time", None),
                "hours": total_hours,
                "rate": hourly_rate,
                "revenue": revenue,
                "revenue_percent": rev_pct,
                "pay_model": pay_model_value,
                "status": safe_text(getattr(shift, "status", None), "pending"),
                "payout": payout,
            }
        )


    wb = openpyxl.Workbook()
    ws_shifts = wb.active
    ws_shifts.title = "Смены"
    ws_summary = wb.create_sheet("Сводка")

    shift_headers = [
        "Сотрудник",
        "Точка",
        "Должность",
        "Дата",
        "Начало смены",
        "Конец смены",
        "Часы",
        "Ставка",
        "Выручка",
        "Процент",
        "Модель оплаты",
        "Статус",
        "К выплате",
    ]
    summary_headers = [
        "Сотрудник",
        "Смен",
        "Часы",
        "Начислено за смены",
        "Бонусы",
        "Штрафы",
        "К выплате",
    ]
    russian_months = [
        "Январь",
        "Февраль",
        "Март",
        "Апрель",
        "Май",
        "Июнь",
        "Июль",
        "Август",
        "Сентябрь",
        "Октябрь",
        "Ноябрь",
        "Декабрь",
    ]

    def style_title(sheet, columns: int) -> None:
        last_col = get_column_letter(columns)
        sheet.merge_cells(f"A1:{last_col}1")
        sheet.merge_cells(f"A2:{last_col}2")
        sheet["A1"] = "Порядок.Смены — отчёт по выплатам"
        sheet["A2"] = f"Период: {russian_months[m - 1]} {y} | Точка: {venue_text}"
        sheet["A1"].font = Font(bold=True, size=14, color="1F2937")
        sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
        sheet["A2"].font = Font(size=11, color="4B5563")
        sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
        sheet["A1"].fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
        sheet["A2"].fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    def apply_header(sheet, row_number: int, headers: list[str]) -> None:
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row_number, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    def set_row_formats(sheet, row_number: int, mapping: dict[int, str]) -> None:
        for col, fmt in mapping.items():
            sheet.cell(row=row_number, column=col).number_format = fmt

    def autofit(sheet, start_row: int, widths: dict[int, int]) -> None:
        for col_idx, base_width in widths.items():
            max_len = base_width
            for row_cells in sheet.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
                value = row_cells[0].value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    if not shift_rows:
        for sheet, columns in ((ws_shifts, len(shift_headers)), (ws_summary, len(summary_headers))):
            style_title(sheet, columns)
            last_col = get_column_letter(columns)
            sheet.merge_cells(f"A4:{last_col}4")
            sheet["A4"] = "За выбранный период смен нет"
            sheet["A4"].font = Font(italic=True, color="6B7280")
            sheet["A4"].alignment = Alignment(horizontal="center", vertical="center")
            sheet.freeze_panes = "A4"
            sheet.sheet_view.showGridLines = False
            autofit(sheet, start_row=1, widths={1: 24, 2: 18, 3: 18, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14, 9: 14, 10: 12, 11: 20, 12: 16, 13: 16})
    else:
        style_title(ws_shifts, len(shift_headers))
        style_title(ws_summary, len(summary_headers))
        ws_shifts.freeze_panes = "A4"
        ws_summary.freeze_panes = "A4"
        ws_shifts.sheet_view.showGridLines = False
        ws_summary.sheet_view.showGridLines = False

        apply_header(ws_shifts, 3, shift_headers)
        ws_shifts.auto_filter.ref = f"A3:M{3 + len(shift_rows)}"

        row = 4
        for item in shift_rows:
            ws_shifts.cell(row=row, column=1, value=item["employee"])
            ws_shifts.cell(row=row, column=2, value=item["venue"])
            ws_shifts.cell(row=row, column=3, value=item["position"])
            ws_shifts.cell(row=row, column=4, value=item["date"])
            ws_shifts.cell(row=row, column=5, value=item["start_time"])
            ws_shifts.cell(row=row, column=6, value=item["end_time"])
            ws_shifts.cell(row=row, column=7, value=float(item["hours"]))
            ws_shifts.cell(row=row, column=8, value=float(item["rate"]))
            ws_shifts.cell(
                row=row,
                column=9,
                value=float(item["revenue"]) if item["revenue"] != Decimal("0.00") else None,
            )
            ws_shifts.cell(
                row=row,
                column=10,
                value=float(item["revenue_percent"]) if item["revenue_percent"] != Decimal("0.00") else None,
            )
            ws_shifts.cell(row=row, column=11, value=pay_model_label(item["pay_model"]))
            ws_shifts.cell(row=row, column=12, value=status_label(item["status"]))
            ws_shifts.cell(row=row, column=13, value=float(item["payout"]))
            set_row_formats(
                ws_shifts,
                row,
                {
                    4: "dd.mm.yyyy",
                    5: "hh:mm",
                    6: "hh:mm",
                    7: "0.00",
                    8: '#,##0.00 ₽',
                    9: '#,##0.00 ₽',
                    10: '0.00',
                    13: '#,##0.00 ₽',
                },
            )
            row += 1

        apply_header(ws_summary, 3, summary_headers)
        ws_summary.auto_filter.ref = f"A3:G{3 + len(user_totals) + 1}"

        row = 4
        total_shifts = 0
        total_hours = Decimal("0.00")
        total_shift_pay = Decimal("0.00")
        total_bonuses = Decimal("0.00")
        total_penalties = Decimal("0.00")
        for totals in sorted(user_totals.values(), key=lambda item: str(item["name"]).lower()):
            net = calculate_payout_total(
                totals["shift_pay"], totals["bonuses"], totals["penalties"]
            )
            total_shifts += int(totals["shifts_count"])
            total_hours += totals["hours"]
            total_shift_pay += totals["shift_pay"]
            total_bonuses += totals["bonuses"]
            total_penalties += totals["penalties"]
            ws_summary.cell(row=row, column=1, value=totals["name"])
            ws_summary.cell(row=row, column=2, value=int(totals["shifts_count"]))
            ws_summary.cell(row=row, column=3, value=float(totals["hours"]))
            ws_summary.cell(row=row, column=4, value=float(totals["shift_pay"]))
            ws_summary.cell(row=row, column=5, value=float(totals["bonuses"]))
            ws_summary.cell(row=row, column=6, value=float(totals["penalties"]))
            ws_summary.cell(row=row, column=7, value=float(net.quantize(Decimal("0.01"))))
            set_row_formats(
                ws_summary,
                row,
                {
                    2: "0",
                    3: "0.00",
                    4: '#,##0.00 ₽',
                    5: '#,##0.00 ₽',
                    6: '#,##0.00 ₽',
                    7: '#,##0.00 ₽',
                },
            )
            row += 1

        ws_summary.cell(row=row, column=1, value="Итого")
        ws_summary.cell(row=row, column=2, value=total_shifts)
        ws_summary.cell(row=row, column=3, value=float(total_hours))
        ws_summary.cell(row=row, column=4, value=float(total_shift_pay))
        ws_summary.cell(row=row, column=5, value=float(total_bonuses))
        ws_summary.cell(row=row, column=6, value=float(total_penalties))
        ws_summary.cell(
            row=row,
            column=7,
            value=float(
                calculate_payout_total(
                    total_shift_pay, total_bonuses, total_penalties
                )
            ),
        )
        for col in range(1, 8):
            cell = ws_summary.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
        set_row_formats(
            ws_summary,
            row,
            {
                2: "0",
                3: "0.00",
                4: '#,##0.00 ₽',
                5: '#,##0.00 ₽',
                6: '#,##0.00 ₽',
                7: '#,##0.00 ₽',
            },
        )

        autofit(
            ws_shifts,
            start_row=3,
            widths={1: 22, 2: 18, 3: 18, 4: 14, 5: 14, 6: 14, 7: 10, 8: 14, 9: 14, 10: 10, 11: 20, 12: 16, 13: 16},
        )
        autofit(
            ws_summary,
            start_row=3,
            widths={1: 22, 2: 10, 3: 10, 4: 16, 5: 12, 6: 12, 7: 14},
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    ascii_filename = f"payroll-{y}-{m:02d}.xlsx"
    russian_filename = f"Порядок.Смены — отчёт по выплатам {russian_months[m - 1].lower()} {y}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{ascii_filename}"; '
                f"filename*=UTF-8''{quote(russian_filename)}"
            )
        },
    )

@router.post("/reminders/shifts")
async def send_shift_reminders(
    session: AsyncSession = Depends(get_session),
    user: User = Depends(get_current_user),
):
    """Send reminders to users who haven't logged today's shift. Call via cron at 21:00."""
    if not _can_send_shift_reminders(user):
        raise HTTPException(
            status_code=403,
            detail="Отправка напоминаний доступна только пользователям с правом управления командой.",
        )

    from app.notifications import send_shift_reminder
    from app.models import Shift as ShiftModel

    today = date.today()

    # Find all active users in all venues
    users_result = await session.execute(
        select(User).where(User.is_active == True, User.telegram_id.isnot(None))
    )
    all_users = users_result.scalars().all()

    # Find users who already have a shift for today
    shifts_result = await session.execute(
        select(ShiftModel.user_id).where(ShiftModel.date == today)
    )
    users_with_shift = {row[0] for row in shifts_result.all()}

    # Send reminders to users without a shift
    reminded = 0
    for user in all_users:
        if user.id not in users_with_shift and user.telegram_id:
            await send_shift_reminder(user.telegram_id)
            reminded += 1

    return {"reminded": reminded}
