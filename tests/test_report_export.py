import ast
import io
import sys
import unittest
import uuid
from datetime import date, datetime, time, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
BACKEND_DIR = REPO_ROOT / "backend"
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from app.models import (  # noqa: E402
    Adjustment,
    AdjustmentType,
    PayModel,
    PayrollRun,
    PayrollRunItem,
    PayrollRunStatus,
    Shift,
    User,
    UserRole,
    Venue,
)
from app.services.report_export import (  # noqa: E402
    MONEY_FORMAT,
    assemble_report_data,
    build_csv,
    build_xlsx,
    safe_spreadsheet_text,
)


API_PATH = BACKEND_DIR / "app" / "routers" / "api.py"
SERVICE_PATH = BACKEND_DIR / "app" / "services" / "report_export.py"
API_SOURCE = API_PATH.read_text(encoding="utf-8")
API_MODULE = ast.parse(API_SOURCE)


def _endpoint_source(name: str) -> str:
    endpoint = next(
        node
        for node in API_MODULE.body
        if isinstance(node, ast.AsyncFunctionDef) and node.name == name
    )
    return ast.get_source_segment(API_SOURCE, endpoint) or ""


def _sample_report():
    home = Venue(id=uuid.uuid4(), name="Главная")
    other = Venue(id=uuid.uuid4(), name="Выездная")
    creator = User(
        id=uuid.uuid4(),
        telegram_id=1,
        name="Администратор",
        role=UserRole.admin,
        venue_id=home.id,
        hourly_rate=Decimal("0"),
        revenue_percentage=Decimal("0"),
        pay_model=PayModel.hourly,
        permissions={},
        is_active=True,
    )
    employee = User(
        id=uuid.uuid4(),
        telegram_id=2,
        name="=ОПАСНАЯ ФОРМУЛА",
        position="Бариста",
        role=UserRole.barista,
        venue_id=home.id,
        hourly_rate=Decimal("9999"),
        revenue_percentage=Decimal("0"),
        pay_model=PayModel.hourly,
        permissions={},
        is_active=True,
    )
    employee.venue = home
    creator.venue = home

    def make_shift(status: str, salary: str, work_venue: Venue, day: int) -> Shift:
        shift = Shift(
            id=uuid.uuid4(),
            user_id=employee.id,
            venue_id=work_venue.id,
            date=date(2026, 7, day),
            start_time=time(9, 0),
            end_time=time(17, 0),
            total_hours=Decimal("8"),
            salary_earned=Decimal(salary),
            revenue=Decimal("10000"),
            status=status,
        )
        shift.user = employee
        shift.venue = work_venue
        return shift

    shifts = [
        make_shift("approved", "1500", other, 2),
        make_shift("pending", "9000", other, 3),
        make_shift("rejected", "8000", home, 4),
    ]

    bonus = Adjustment(
        id=uuid.uuid4(),
        user_id=employee.id,
        venue_id=other.id,
        type=AdjustmentType.bonus,
        amount=Decimal("200"),
        reason="=SUM(A1:A2)",
        created_by=creator.id,
        month=7,
        year=2026,
        created_at=datetime(2026, 7, 5, 12, tzinfo=timezone.utc),
    )
    deduction = Adjustment(
        id=uuid.uuid4(),
        user_id=employee.id,
        venue_id=other.id,
        type=AdjustmentType.penalty,
        amount=Decimal("50"),
        reason="Аванс",
        created_by=creator.id,
        month=7,
        year=2026,
        created_at=datetime(2026, 7, 6, 12, tzinfo=timezone.utc),
    )
    for adjustment in (bonus, deduction):
        adjustment.user = employee
        adjustment.venue = other
        adjustment.creator = creator

    payroll_run = PayrollRun(
        id=uuid.uuid4(),
        title="Июльский расчёт",
        period_start=date(2026, 7, 1),
        period_end=date(2026, 7, 15),
        status=PayrollRunStatus.finalized,
        total_amount=Decimal("99999"),
        total_paid=Decimal("99999"),
        revenue_total=Decimal("12000"),
        created_by_id=creator.id,
        venue_id=other.id,
        created_at=datetime(2026, 7, 16, 10, tzinfo=timezone.utc),
        finalized_at=datetime(2026, 7, 16, 11, tzinfo=timezone.utc),
    )
    payroll_run.venue = other
    item = PayrollRunItem(
        id=uuid.uuid4(),
        payroll_run_id=payroll_run.id,
        user_id=employee.id,
        approved_shifts_count=1,
        approved_hours=Decimal("8"),
        base_amount=Decimal("1500"),
        bonus_amount=Decimal("200"),
        deduction_amount=Decimal("50"),
        final_amount=Decimal("1650"),
        paid_amount=Decimal("500"),
        remaining_amount=Decimal("1150"),
    )
    payroll_run.items = [item]

    return assemble_report_data(
        month=7,
        year=2026,
        venue_name="Выездная",
        shifts=shifts,
        adjustments=[bonus, deduction],
        payroll_runs=[payroll_run],
    )


class ReportExportTests(unittest.TestCase):
    def test_permissions_and_actual_venue_scope_are_kept_in_router(self) -> None:
        helper = _endpoint_source("_get_export_report")
        self.assertIn('has_permission(user, "can_export_payroll")', helper)
        self.assertIn("status_code=403", helper)
        scope_node = next(
            node for node in API_MODULE.body
            if isinstance(node, ast.FunctionDef) and node.name == "_export_scope_venue_id"
        )
        scope_source = ast.get_source_segment(API_SOURCE, scope_node) or ""
        self.assertIn("user.venue_id", scope_source)
        service_source = SERVICE_PATH.read_text(encoding="utf-8")
        self.assertIn("Shift.venue_id == venue_id", service_source)
        self.assertIn("Adjustment.venue_id == venue_id", service_source)
        self.assertIn("PayrollRun.venue_id == venue_id", service_source)

    def test_all_statuses_are_exported_but_financial_totals_use_approved_only(self) -> None:
        report = _sample_report()
        self.assertEqual(len(report.shifts), 3)
        self.assertEqual({row["status"] for row in report.shifts}, {"Утверждена", "На подтверждении", "Отклонена"})
        employee = report.employees[0]
        self.assertEqual(employee["approved_shifts"], 1)
        self.assertEqual(employee["shift_amount"], Decimal("1500"))
        self.assertEqual(employee["total"], Decimal("1650"))

    def test_historical_salary_and_home_actual_venue_are_not_mixed(self) -> None:
        report = _sample_report()
        approved = next(row for row in report.shifts if row["is_approved"])
        self.assertEqual(approved["salary"], Decimal("1500"))
        self.assertEqual(approved["home_venue"], "Главная")
        self.assertEqual(approved["work_venue"], "Выездная")
        self.assertEqual(report.employees[0]["cross_venue"], 1)
        overview = {label: value for label, value, _ in report.overview}
        self.assertEqual(overview["Cross-venue смены"], 1)

    def test_overview_separates_shift_accruals_and_adjustments(self) -> None:
        report = _sample_report()
        overview = {label: value for label, value, _ in report.overview}
        self.assertEqual(overview["Начислено за смены"], Decimal("1500"))
        self.assertEqual(overview["Бонусы"], Decimal("200"))
        self.assertEqual(overview["Удержания"], Decimal("50"))
        self.assertEqual(overview["Итого к выплате"], Decimal("1650"))

    def test_payroll_totals_use_item_snapshots(self) -> None:
        report = _sample_report()
        run = report.payroll_runs[0]
        self.assertEqual(run["accrued"], Decimal("1650"))
        self.assertEqual(run["paid"], Decimal("500"))
        self.assertEqual(run["remaining"], Decimal("1150"))
        overview = {label: value for label, value, _ in report.overview}
        self.assertEqual(overview["Зафиксировано"], Decimal("1650"))
        self.assertEqual(overview["Выплачено"], Decimal("500"))

    def test_workbook_structure_formatting_and_empty_period(self) -> None:
        report = _sample_report()
        workbook = load_workbook(io.BytesIO(build_xlsx(report)))
        self.assertEqual(
            workbook.sheetnames,
            ["Обзор", "Смены", "Сотрудники", "Корректировки", "Расчёты и выплаты"],
        )
        for sheet in workbook.worksheets:
            self.assertEqual(sheet.freeze_panes, "A5")
            self.assertFalse(sheet.sheet_view.showGridLines)
            self.assertTrue(sheet.auto_filter.ref)
        self.assertIn("ShiftsTable", workbook["Смены"].tables)
        self.assertEqual(workbook["Смены"]["K4"].value, "Текущая ставка")
        self.assertEqual(workbook["Смены"]["M5"].number_format, MONEY_FORMAT)
        self.assertEqual(workbook["Расчёты и выплаты"]["J5"].number_format, "0.00%")

        empty = assemble_report_data(
            month=7, year=2026, venue_name="Все точки", shifts=[], adjustments=[], payroll_runs=[]
        )
        empty_book = load_workbook(io.BytesIO(build_xlsx(empty)))
        self.assertEqual(empty_book["Смены"]["A5"].value, "За выбранный период смен нет")

    def test_formula_injection_is_neutralized_in_xlsx_and_csv(self) -> None:
        self.assertEqual(safe_spreadsheet_text("=1+1"), "'=1+1")
        report = _sample_report()
        workbook = load_workbook(io.BytesIO(build_xlsx(report)), data_only=False)
        self.assertEqual(workbook["Смены"]["B5"].value, "'=ОПАСНАЯ ФОРМУЛА")
        self.assertEqual(workbook["Корректировки"]["F5"].value, "'=SUM(A1:A2)")
        csv_text = build_csv(report).decode("utf-8-sig")
        self.assertTrue(build_csv(report).startswith(b"\xef\xbb\xbf"))
        self.assertIn(";", csv_text.splitlines()[0])
        self.assertIn("'=ОПАСНАЯ ФОРМУЛА", csv_text)

    def test_filename_headers_have_ascii_and_utf8_names(self) -> None:
        source = API_SOURCE
        self.assertIn('filename="{ascii_filename}"', source)
        self.assertIn("filename*=UTF-8''", source)
        self.assertIn('report.month, report.year, "xlsx"', _endpoint_source("export_xlsx"))
        self.assertIn('report.month, report.year, "csv"', _endpoint_source("export_csv"))


if __name__ == "__main__":
    unittest.main()
